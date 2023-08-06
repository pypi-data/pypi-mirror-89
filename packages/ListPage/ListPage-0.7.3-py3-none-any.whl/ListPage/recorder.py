#!/usr/bin/env python
# -*- coding:utf-8 -*-

import csv
from pathlib import Path
from typing import Union


class Recorder(object):
    """Recorder对象用于暂缓写入数据，
    它可接收列表数据，达到一定数量时才一次进行写入，
    以降低文件读写次数，减少开销。"""

    def __init__(self, file_path: str = None, cache_size: int = 50):
        """初始化                                                    \n
        :param file_path: 保存的文件路径
        :param cache_size: 每接收多少条记录写入文件
        """
        self.cache_size = cache_size
        self._data = []
        self._before = []
        self._after = []
        self._file_path = None
        self.encoding = 'utf-8'

        self.file_path = file_path

    def __del__(self) -> None:
        """对象关闭时把剩下的数据写入文件"""
        self.record()

    @property
    def cache_size(self) -> int:
        """返回缓存大小"""
        return self._cache

    @cache_size.setter
    def cache_size(self, cache_size: int) -> None:
        """设置缓存大小                   \n
        :param cache_size: 缓存大小
        :return: None
        """
        if not isinstance(cache_size, int):
            raise TypeError('参数cache_size只能为int。')

        self._cache = cache_size

    @property
    def file_path(self) -> str:
        """返回文件路径"""
        if not self._file_path:
            raise ValueError('未指定文件路径。')

        return self._file_path

    @file_path.setter
    def file_path(self, file_path: str) -> None:
        """设置文件路径                \n
        :param file_path: 文件路径
        :return: None
        """
        if not isinstance(file_path, str):
            raise TypeError(f'参数file_path只能是str，非{type(file_path)}')

        elif not file_path.endswith(('.xlsx', '.csv', '.json', '.txt')):
            raise TypeError('只支持xlsx、csv、txt和json格式文件。')

        if self._file_path:
            self.record()

        self._file_path = file_path

    def add_data(self, data: Union[list, tuple]) -> None:
        """添加数据                                                           \n
        :param data: 插入的数据，元组或列表
        :return: None
        """
        if isinstance(data, (list, tuple)) and not isinstance(data[0], (list, tuple)):
            self._data.append(data)
        else:
            self._data.extend(data)

        if len(self._data) >= self.cache_size:
            self.record()

    def record(self) -> None:
        """记录数据"""
        if not self._data:
            return

        file_type = self.file_path.split('.')[-1]

        if file_type == 'xlsx':
            _record_to_xlsx(self.file_path, self._data, self._before, self._after)

        elif file_type == 'csv':
            _record_to_csv(self.file_path, self._data, self._before, self._after, self.encoding)

        elif file_type == 'txt':
            _record_to_txt(self.file_path, self._data, self._before, self._after, self.encoding)

        elif file_type == 'json':
            _record_to_json(self.file_path, self._data, self._before, self._after, self.encoding)

        self._data = []

    def set_head(self, head: Union[list, tuple]) -> None:
        """设置表头。只有 csv 和 xlsx 格式支持设置表头                           \n
        :param head: 表头，列表或元组
        :return: None
        """
        file_type = self.file_path.split('.')[-1]

        if file_type == 'xlsx':
            _set_xlsx_head(self.file_path, head)

        elif file_type == 'csv':
            _set_csv_head(self.file_path, head, self.encoding)

        else:
            print('只能对xlsx和csv文件设置表头。')

    def set_before(self, before: Union[list, tuple, str, dict]) -> None:
        """设置在数据前面补充的列                                                    \n
        :param before: 列表、元组或字符串，为字符串时则补充一列
        :return: None
        """
        self.record()

        if before is None:
            self._before = []
        elif isinstance(before, (list, dict)):
            self._before = before
        elif isinstance(before, tuple):
            self._before = list(before)
        else:
            self._before = [str(before)]

    def set_after(self, after: Union[list, tuple, str, dict]) -> None:
        """设置在数据后面补充的列                                                      \n
        :param after: 列表、元组或字符串，为字符串时则补充一列
        :return: None
        """
        self.record()

        if after is None:
            self._after = []
        elif isinstance(after, (list, dict)):
            self._after = after
        elif isinstance(after, tuple):
            self._after = list(after)
        else:
            self._before = [str(after)]

    def clear(self) -> None:
        """清空缓存中的数据"""
        self._data = []


def _record_to_xlsx(file_path: str,
                    data: list,
                    before: Union[list, dict] = None,
                    after: Union[list, dict] = None) -> None:
    """记录数据到xlsx文件            \n
    :param file_path: 文件路径
    :param data: 要记录的数据
    :param before: 数据前面的列
    :param after: 数据后面的列
    :return:
    """
    try:
        import openpyxl

    except ModuleNotFoundError:
        import os
        os.system('pip install -i https://pypi.tuna.tsinghua.edu.cn/simple openpyxl')
        import openpyxl

    if Path(file_path).exists():
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

    else:
        wb = openpyxl.Workbook()
        ws = wb.active

        title = _get_title(data[0], before, after)
        if title is not None:
            ws.append(title)

    for i in data:
        ws.append(_data_to_list(i, before, after))

    wb.save(file_path)
    wb.close()


def _record_to_csv(file_path: str,
                   data: list,
                   before: Union[list, dict] = None,
                   after: Union[list, dict] = None,
                   encoding: str = 'utf-8') -> None:
    """记录数据到csv文件            \n
    :param file_path: 文件路径
    :param data: 要记录的数据
    :param before: 数据前面的列
    :param after: 数据后面的列
    :param encoding: 编码
    :return:
    """

    exists = True if Path(file_path).exists() else False

    with open(file_path, 'a+', newline='', encoding=encoding) as f:
        csv_write = csv.writer(f)

        if not exists:
            title = _get_title(data[0], before, after)
            if title:
                csv_write.writerow(title)

        for i in data:
            csv_write.writerow(_data_to_list(i, before, after))


def _record_to_txt(file_path: str,
                   data: list,
                   before: Union[list, dict] = None,
                   after: Union[list, dict] = None,
                   encoding: str = 'utf-8') -> None:
    """记录数据到txt文件            \n
    :param file_path: 文件路径
    :param data: 要记录的数据
    :param before: 数据前面的列
    :param after: 数据后面的列
    :param encoding: 编码
    :return:
    """
    with open(file_path, 'a+', encoding=encoding) as f:
        for i in data:
            i = _data_to_list_or_dict(i, before, after)
            f.write(f'{str(i)}\n')


def _record_to_json(file_path: str,
                    data: list,
                    before: Union[list, dict] = None,
                    after: Union[list, dict] = None,
                    encoding: str = 'utf-8') -> None:
    """记录数据到json文件            \n
    :param file_path: 文件路径
    :param data: 要记录的数据
    :param before: 数据前面的列
    :param after: 数据后面的列
    :param encoding: 编码
    :return:
    """
    import json

    if Path(file_path).exists():
        with open(file_path, 'r', encoding=encoding) as f:
            json_data = json.load(f)

        for i in data:
            i = _data_to_list_or_dict(i, before, after)
            json_data.append(i)

    else:
        json_data = [_data_to_list_or_dict(i, before, after) for i in data]

    with open(file_path, 'w', encoding=encoding) as f:
        json.dump(json_data, f)


def _set_xlsx_head(file_path: str, head: Union[list, tuple]) -> None:
    """设置xlsx文件的表头                       \n
    :param file_path: 文件路径
    :param head: 表头列表或元组
    :return: None
    """
    try:
        import openpyxl
    except ModuleNotFoundError:
        import os
        os.system('pip install -i https://pypi.tuna.tsinghua.edu.cn/simple openpyxl')
        import openpyxl

    if Path(file_path).exists():
        wb = openpyxl.load_workbook(file_path)
    else:
        wb = openpyxl.Workbook()

    ws = wb.active

    for key, i in enumerate(head, 1):
        ws.cell(1, key).value = i

    wb.save(file_path)
    wb.close()


def _set_csv_head(file_path: str, head: Union[list, tuple], encoding: str = 'utf-8') -> None:
    """设置csv文件的表头                       \n
    :param file_path: 文件路径
    :param head: 表头列表或元组
    :param encoding: 编码
    :return: None
    """
    head_txt = ','.join(head)

    if Path(file_path).exists():
        with open(file_path, 'r', newline='', encoding=encoding) as f:
            content = "".join(f.readlines()[1:])

        with open(file_path, 'w', newline='', encoding=encoding) as f:
            f.write(f'{head_txt}\n{content}')

    else:
        with open(file_path, 'w', newline='', encoding=encoding) as f:
            f.write(f'{head_txt}\n')


def _get_title(data: Union[list, dict],
               before: Union[list, dict, None] = None,
               after: Union[list, dict, None] = None) -> Union[list, None]:
    """获取表头列表                         \n
    :param data: 数据列表或字典
    :param before: 数据前的列
    :param after: 数据后的列
    :return: 表头列表
    """
    if isinstance(data, (tuple, list)):
        return None

    return_list = []

    for i in (before, data, after):
        if isinstance(i, dict):
            return_list.extend(list(i))
        elif i is None:
            pass
        elif isinstance(i, list):
            return_list.extend(['' for _ in range(len(i))])
        else:
            return_list.extend([''])

    return return_list


def _data_to_list(data: Union[list, dict],
                  before: Union[list, dict, None] = None,
                  after: Union[list, dict, None] = None) -> list:
    """将传入的数据转换为列表形式          \n
    :param data: 要处理的数据
    :return: 转变成列表方式的数据
    """
    return_list = []

    for i in (before, data, after):
        if isinstance(i, dict):
            return_list.extend(list(i.values()))
        elif i is None:
            pass
        elif isinstance(i, list):
            return_list.extend(i)
        elif isinstance(i, tuple):
            return_list.extend(list(i))
        else:
            return_list.extend([str(i)])

    return return_list


def _data_to_list_or_dict(data: Union[list, dict],
                          before: Union[list, dict, None] = None,
                          after: Union[list, dict, None] = None) -> Union[list, dict]:
    """将传入的数据转换为列表或字典形式          \n
    :param data: 要处理的数据
    :return: 转变成列表方式的数据
    """
    if isinstance(data, list):
        return _data_to_list(data, before, after)

    elif isinstance(data, dict):
        if isinstance(before, dict):
            data = {**before, **data}

        if isinstance(after, dict):
            data = {**data, **after}

        return data
