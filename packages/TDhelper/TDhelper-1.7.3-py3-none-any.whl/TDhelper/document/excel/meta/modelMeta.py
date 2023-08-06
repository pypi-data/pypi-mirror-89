from TDhelper.generic.transformationType import transformation
from TDhelper.document.excel.FieldType import *
from types import FunctionType, MethodType, ModuleType
from openpyxl import load_workbook
import csv
import copy


class _AttributeOverride:
    def __init__(self, name, m_type):
        self._name = name
        self._type = m_type

    def __get__(self, instance, owen):
        return instance.__dict__[self._name]

    def __set__(self, instance, value):
        instance.__dict__[self._name] = transformation(value, self._type)

    def __delete__(self, instance):
        instance.__dict__.pop(self._name)


class Meta:
    file = None
    sheet = 'sheet1'
    extension = 'xlsx'


class modelMeta(type):
    def __new__(cls, name, bases, dct):
        attrs = {'mapping': {}, 'Meta': Meta, '__exit__': __exit__, 'readLine': readLine,
                 'close': close, '__initExcelHandle__': __initExcelHandle__, 'rows': rows, 'getCount': getCount, '_translateMode': _translateMode, '__rowsCount__': 0}
        for name, value in dct.items():
            if (not isinstance(dct[name], type)) and (not isinstance(dct[name], FunctionType)):
                if not name.startswith('__'):
                    if isinstance(dct[name], FieldType):
                        attrs['mapping'][name] = value.bindCol
                        attrs[name] = _AttributeOverride(name, value.fieldType)
                    else:
                        raise Exception('field type must is FieldType.')
                else:
                    attrs[name] = value
            else:
                if isinstance(dct[name], type):
                    if name == 'Meta':
                        for attr_name in dct[name].__dict__:
                            if not attr_name.startswith('__'):
                                setattr(attrs['Meta'], attr_name,
                                        dct[name].__dict__[attr_name])
                else:
                    attrs[name] = value
        return super(modelMeta, cls).__new__(cls, name, bases, attrs)


def __initExcelHandle__(self):
    try:
        if self.Meta.file:
            m_extension = self.Meta.file.rsplit('.')[1]
            if m_extension == 'csv':
                self.Meta.extension = 'csv'
                self.__excelHandle__ = open(self.Meta.file)
                self.__sheetHandle__ = csv.reader(self.__excelHandle__)
                self.__excelHandle__ = None
                m_offset = 1
                for v in self.__sheetHandle__:
                    if v:
                        if m_offset < self.__record_offset__:
                            m_offset += 1
                            continue
                        self.__rows__.append(self._translateMode(v))
                    else:
                        break
                self.__rowsCount__ = len(self.__rows__)
            elif m_extension == 'xlsx' or m_extension == 'xls':
                self.Meta.extension = 'xlsx'
                self.__excelHandle__ = load_workbook(self.Meta.file)
                self.__sheetHandle__ = self.__excelHandle__[self.Meta.sheet]
                column = self.__sheetHandle__.max_column+1
                m_offset= 1
                tmp_handle= self.__sheetHandle__
                for m_rows in tmp_handle:
                    if m_offset < self.__record_offset__:
                        m_offset += 1
                        continue
                    self.__rows__.append(self._translateMode(m_rows))
                self.__rowsCount__ = len(self.__rows__)
                tmp_handle= None
            else:
                raise Exception('file extension is error.')
        else:
            raise Exception('meta file is None.')
    except Exception as e:
        raise e


def _translateMode(self, data=[]):
    m_extension = self.Meta.file.rsplit('.')[1].lower()
    if data:
        try:
            # 关闭IO，然后返回深拷贝对象解决数组里对象值引用引起的所有对象同样的值.
            # (没搞明白，为什么这里设为NONE了。外层的for循环还能继续下去)
            self.__sheetHandle__ = None
            self.__excelHandle__ = None
            for (name, value) in self.mapping.items():
                if value <= len(data):
                    if m_extension == 'csv':
                        attr_value = data[value-1]
                    elif m_extension == 'xlsx' or m_extension == 'xls':
                        attr_value= data[value-1].value
                    else:
                        raise Exception('file extension is error.')
                    if not attr_value:
                        attr_value = '0.00'
                    if attr_value == 'None':
                        attr_value = '0.00'
                    setattr(self, name, attr_value)
                else:
                    raise Exception('mapping error:(%s,%s)' % (name, value))
            return copy.deepcopy(self)
        except Exception as e:
            raise e
    else:
        raise Exception("translate data is none.")


def __exit__(self, exc_type, exc_value, exc_t):
    self.close()


def close(self):
    self.__excelHandle__ = None
    self.__sheetHandle__ = None


def getCount(self):
    return self.__rowsCount__


def rows(self):
    return self.__rows__


def readLine(self, lineOffset=0):
    if lineOffset >= self.getCount():
        return None
    self.__read_offset__= lineOffset
    return self.__rows__[lineOffset]
