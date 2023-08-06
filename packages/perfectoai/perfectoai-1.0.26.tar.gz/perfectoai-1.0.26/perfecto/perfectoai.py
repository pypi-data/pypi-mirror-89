#!/usr/bin/env python
import argparse
import base64
import collections
import configparser
import datetime
import glob
import html
import io
import json
import multiprocessing
import os
import platform
import re
import shutil
import ssl
import sys
import tempfile
import time
import traceback
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
import uuid
import webbrowser
import xml.etree.ElementTree as ETree
from collections import Counter
from datetime import datetime, timedelta
from multiprocessing import Pool, Process, freeze_support
from re import MULTILINE
from urllib.error import HTTPError
from xml.dom import minidom

import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import numpy as np
import pandas
import plotly
import plotly.express as px
import pylab as pl
import requests
import tzlocal
from colorama import init
from dateutil.parser import parse
from easydict import EasyDict as edict
from fbprophet.plot import plot_plotly
from jenkinsapi.jenkins import Jenkins
from json2html import *
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment
from pandas.io.json import json_normalize
from pandas.plotting import table
from termcolor import colored

""" Microsoft Visual C++ required, cython required for pandas installation, """
TEMP_DIR = "/tmp" if platform.system() == "Darwin" else tempfile.gettempdir()
live_report_filename = "Interactive_Report.html"
email_report_filename = "email.html"
orchestrationIssues = ["already in use"]
labIssues = ["HANDSET_ERROR", "ERROR: No device was found"]
# Do not change these variable
RESOURCE_TYPE = "handsets"
RESOURCE_TYPE_USERS = "users"
REPOSITORY_RESOURCE_TYPE = "repositories/media"
report = "Report: "
tags = ""
reportTag = ""
criteria = ""
jobName = ""
jobNumber = ""
startDate = ""
endDate = ""
consolidate = ""
trends = "false"
port = ""
temp = ""
resources = []
topTCFailureDict = {}
topDeviceFailureDict = {}
labIssuesCount = 0
scriptingIssuesCount = 0
orchestrationIssuesCount = 0
cleanedFailureList = {}
suggesstionsDict = {}
failurereasons = {}
issues = {}
issues_email = ""
topfailedtable = {}
monthlyStats = {}
title = ""
execution_summary = {}
recommendations = {}
execution_status = {}


def send_request(url):
    """send request"""
    #     print("Submitting", url)
    device_list_parameters = os.environ["DEVICE_LIST_PARAMETERS"]
    if (
        "All devices" in device_list_parameters
        or "Available devices only" in device_list_parameters
    ):
        response = urllib.request.urlopen(url)
    else:
        response = urllib.request.urlopen(url.replace(" ", "%20"))
    #    rc = response.getcode()
    #    print("rc =", rc)
    return response


def send_request_with_json_response(url):
    """send request"""
    response = send_request(url)
    text = response.read().decode("utf-8")
    maps = json.loads(text)
    return maps


"""
   returns as text if none
"""


def as_text(value):
    """as texts"""
    if value is None:
        return ""
    return str(value)


def convertxmlToXls(xml, dict_keys, filename):
    """
            Checks if file exists, parses the file and extracts the needed data
            returns a 2 dimensional list without "header"
    """
    root = ETree.fromstring(xml)
    headers = []
    finalHeaders = []
    if dict_keys is None:
        for child in root:
            headers.append({x.tag for x in root.findall(child.tag + "/*")})
    else:
        headers = dict_keys
    headers = headers[0]
    mdlist = []
    for child in root:
        temp = []
        for key in sorted(headers):
            try:
                finalHeaders.append(key)
                temp.append(child.find(key).text)
            except Exception:
                temp.append("-")
        mdlist.append(temp)
    """
    Generates excel file with given data
    mdlist: 2 Dimensional list containing data
    """
    wb = Workbook()
    ws = wb.active
    for i, row in enumerate(mdlist):
        for j, value in enumerate(row):
            ws.cell(row=i + 1, column=j + 1).value = value
    ws.insert_rows(0)
    # generates header
    i = 0
    finalHeaders = list(dict.fromkeys(finalHeaders))
    for i, value in enumerate(finalHeaders):
        ws.cell(1, column=i + 1).value = value
        ws.cell(1, column=i + 1).alignment = Alignment(horizontal="center")
    for column_cells in ws.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 5
    newfilename = os.path.abspath(filename)
    wb.save(newfilename)
    return


def user_condition(df):
    cols = list(df)
    if len(cols) > 0:
        cols = [cols[-1]] + cols[:-1]
        df = df[cols]
        df = df.replace(np.nan, "", regex=True)
        df = df[~df["email"].str.contains("perfectomobile.com")]
        df = df.sort_values(by="firstName")
    return df


def convertjsonToXls(json_text, dict_keys, filename):
    jsonfile = "user_results.json"
    file = os.path.join(TEMP_DIR, "results", jsonfile)
    f = open(file, "w+")
    f.write(str(json_text))
    f.close()
    data = json.load(open(file))
    sys.stdout.flush()
    pandas.set_option("display.max_columns", 6)
    # pandas.set_option('display.max_colwidth', 120)
    pandas.set_option("colheader_justify", "left")
    df = pandas.DataFrame(data["users"])
    if len(df.index) < 1:
        raise Exception(
            "There are no users who match the expected conditions "
            + os.environ["USER_LIST_PARAMETERS"]
        )
    else:
        df.drop(
            [
                "username",
                "authentication",
                "gender",
                "phoneNumberExt",
                "location",
                "stateCode",
                "state",
            ],
            axis=1,
            inplace=True,
            errors="ignore",
        )
    df = user_condition(df)
    df.to_excel(filename, index=False)
    wb = Workbook()
    wb = load_workbook(filename)
    ws = wb.worksheets[0]
    for column_cells in ws.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 5
    newfilename = os.path.abspath(filename)
    wb.save(newfilename)
    return df


def send_request_with_xml_response(url):
    """send request"""
    response = send_request(url)
    decoded = response.read().decode("utf-8")
    xmldoc = minidom.parseString(decoded)
    return xmldoc


def send_request_to_xlsx(url, filename):
    """send_request_to_xlsx"""
    response = send_request(url)
    decoded = response.read().decode("utf-8")
    if any(["=list" in url, "=users" in url]):
        filename = os.path.join(TEMP_DIR, "output", filename)
        convertxmlToXls(decoded, None, filename)


def send_jsonrequest_to_xlsx(url, filename):
    """send_request_to_xlsx"""
    try:
        response = send_request(url)
    except:
        raise Exception(
            "unable to find users who match the expected conditions "
            + os.environ["USER_LIST_PARAMETERS"]
        )
    decoded = response.read().decode("utf-8")
    if any(["=list" in url, "=users" in url]):
        filename = os.path.join(TEMP_DIR, "output", filename)
        return convertjsonToXls(decoded, None, filename)


def send_request2(url):
    """send request"""
    response = send_request(url)
    text = response.read().decode("utf-8")
    return text


def get_url(resource, resource_id, operation):
    """get url """
    cloudname = os.environ["cloudName"]
    url = "https://" + cloudname + ".perfectomobile.com/services/" + resource
    if resource_id != "":
        url += "/" + str(resource_id)
    token = os.environ["securityToken"]
    if "eyJhb" in token:
        query = urllib.parse.urlencode({"operation": operation, "securityToken": token})
    else:
        if ":" in token:
            user = token.split(":")[0]
            pwd = token.split(":")[1]
            query = urllib.parse.urlencode(
                {"operation": operation, "user": user, "password": pwd}
            )
        else:
            raise Exception(
                "Please pass your perfecto credentials in the format user:password as -s parameter value. Avoid using special characters such as :,@. in passwords!"
            )
    url += "?" + query
    return url

def getregex_output(response, pattern1, pattern2):
    """regex"""
    matches = re.finditer(pattern1, response, re.MULTILINE)
    for match in matches:
        match_item = str(re.findall(pattern2, match.group()))
        match_item = match_item.replace(':"', "").replace('"', "")
        match_item = match_item.replace("'", "").replace("[", "")
        match_item = (
            match_item.replace("]", "")
            .replace(",test", "")
            .replace(",timer.system", "")
            .replace('description":"', "")
        )
        return str(match_item)

def getregex_error(response, pattern1, pattern2):
    """regex"""
    matches = re.finditer(pattern1, response, re.MULTILINE)
    for match in matches:
        match_item = str(re.findall(pattern2, match.group()))
        match_item = match_item.replace("^'", "").replace("' $", "")
        return str(match_item)


def device_command(exec_id, device_id, operation):
    """Runs device command"""
    url = get_url("executions/" + str(exec_id), "", "command")
    url += "&command=" + "device"
    url += "&subcommand=" + operation
    url += "&param.deviceId=" + device_id
    send_request_with_json_response(url)


def end_execution(exec_id):
    """End execution"""
    url = get_url("executions/" + str(exec_id), "", "end")
    send_request_with_json_response(url)


def start_exec():
    """start execution"""
    url = get_url("executions", "", "start")
    response = send_request2(url)
    exec_id = getregex_output(response, r"executionId\"\:\"[\w\d@.-]+\"", ':".*$')
    return exec_id

def get_device_list_response(resource, command, status, in_use):
    """get_device_list_response"""
    url = get_url(resource, "", command)
    url += "&status=" + status
    if in_use != "":
        url += "&inUse=" + in_use
    if len(os.environ["DEVICE_LIST_PARAMETERS"].split(":")) >= 2:
        for item in os.environ["DEVICE_LIST_PARAMETERS"].split(";"):
            if ":" in item:
                url += "&" + item.split(":")[0] + "=" + item.split(":")[1]
    xmldoc = send_request_with_xml_response(url)
    return xmldoc


def get_xml_to_xlsx(resource, command, filename):
    """get_xml_to_xlsx"""
    url = get_url(resource, "", command)
    send_request_to_xlsx(url, filename)
    sys.stdout.flush()


def get_json_to_xlsx(resource, command, filename):
    """get_json_to_xlsx"""
    url = get_url(resource, "", command)
    if "All users" not in os.environ["USER_LIST_PARAMETERS"]:
        if len(os.environ["USER_LIST_PARAMETERS"].split(":")) >= 2:
            for item in os.environ["USER_LIST_PARAMETERS"].split(";"):
                if ":" in item:
                    url += "&" + item.split(":")[0] + "=" + item.split(":")[1]
    return send_jsonrequest_to_xlsx(url.replace(" ", "%20"), filename)


def get_device_ids(xmldoc):
    """get_device_ids"""
    device_ids = xmldoc.getElementsByTagName("deviceId")
    return device_ids


def get_handset_count(xmldoc):
    """get_handset_count"""
    handset_elements = xmldoc.getElementsByTagName("handset")
    return len(handset_elements)


def exec_command(exec_id, device_id, cmd, subcmd):
    """exec_commands"""
    url = get_url("executions/" + str(exec_id), "", "command")
    url += "&command=" + cmd
    url += "&subcommand=" + subcmd
    url += "&param.deviceId=" + device_id
    response = send_request2(url)
    status = getregex_output(
        response,
        r"(description\"\:\".*\",\"timer.system|returnValue\"\:\".*\",\"test)",
        ':".*$',
    )
    return str(status)


def perform_actions(deviceid_color):
    """perform_actions"""
    get_network_settings = os.environ["GET_NETWORK_SETTINGS"]
    deviceid_color = str(deviceid_color)
    device_id = deviceid_color.split("||", 1)[0]
    color = deviceid_color.split("||", 1)[1]
    desc = deviceid_color.split("||", 2)[2]
    fileName = device_id + ".txt"
    file = os.path.join(TEMP_DIR, "results", fileName)
    try:
        status = "Results="
        # update dictionary
        url = get_url(RESOURCE_TYPE, device_id, "info")
        xmldoc = send_request_with_xml_response(url)
        modelElements = xmldoc.getElementsByTagName("model")
        manufacturerElements = xmldoc.getElementsByTagName("manufacturer")
        model = modelElements[0].firstChild.data
        manufacturer = manufacturerElements[0].firstChild.data
        osElements = xmldoc.getElementsByTagName("os")
        osDevice = osElements[0].firstChild.data
        try:
            descriptionElements = xmldoc.getElementsByTagName("description")
            description = descriptionElements[0].firstChild.data
        except:
            description = ""
        try:
            osVElements = xmldoc.getElementsByTagName("osVersion")
            osVersion = osVElements[0].firstChild.data
        except:
            osVersion = "NA"
        osVersion = osDevice + " " + osVersion
        try:
            operatorElements = xmldoc.getElementsByTagName("operator")
            operator = operatorElements[0].childNodes[0].data
        except:
            operator = "NA"
        try:
            phElements = xmldoc.getElementsByTagName("phoneNumber")
            phoneNumber = phElements[0].firstChild.data
        except:
            phoneNumber = "NA"
        if "green" in color:
            start_execution = os.environ["START_EXECUTION"]
            if "true" in start_execution.lower():
                # Get execution id
                EXEC_ID = start_exec()
                # open device:
                print("opening: " + model + ", device id: " + device_id)
                device_command(EXEC_ID, device_id, "open")
                cleanup = os.environ["CLEANUP"]
                if "True" in cleanup:
                    if not "iOS" in osDevice:
                        print("cleaning up: " + model + ", device id: " + device_id)
                        try:
                            status += "clean:" + str(
                                exec_command(EXEC_ID, device_id, "device", "clean")
                            ).replace(",", " ")
                        except:
                            status += "clean:Failed!"
                        status += ";"
                    else:
                        status += "clean:NA;"
                reboot = os.environ["REBOOT"]
                if "True" in reboot:
                    if all(
                        [
                            "Huawei" not in manufacturer,
                            "Xiaomi" not in manufacturer,
                            "Oppo" not in manufacturer,
                            "Motorola" not in manufacturer,
                            "OnePlus" not in manufacturer,
                        ]
                    ):
                        print("rebooting: " + model + ", device id: " + device_id)
                        try:
                            status += "reboot:" + str(
                                exec_command(EXEC_ID, device_id, "device", "reboot")
                            ).replace(",", " ")
                        except:
                            status += "reboot:Failed!"
                        status += ";"
                    else:
                        print(model + " not applicable for rebooting")
                        status += "reboot:NA;"
                if "True" in get_network_settings:
                    print(
                        "getting network status of : "
                        + model
                        + ", device id: "
                        + device_id
                    )
                    networkstatus = "airplanemode=Failed, wifi=Failed, data=Failed"
                    try:
                        tempstatus = (
                            str(
                                exec_command(
                                    EXEC_ID, device_id, "network.settings", "get"
                                )
                            )
                            .replace("{", "")
                            .replace("}", "")
                        )
                        if tempstatus.count(",") == 2:
                            networkstatus = tempstatus
                            status += "NW:OK"
                        else:
                            status += "NW:Failed!"
                    except:
                        status += "NW:Failed!"
                    status += ";"
                # Close device
                print("closing: " + model + ", device id: " + device_id)
                device_command(EXEC_ID, device_id, "close")
                # End execution
                end_execution(EXEC_ID)
        else:
            networkstatus = ",,"

        if "True" in get_network_settings:
            final_string = (
                "status="
                + desc
                + ", deviceId='"
                + device_id
                + "', Manufacturer="
                + str(manufacturer)
                + "', model="
                + str(model)
                + ", version="
                + str(osVersion)
                + ", description="
                + str(description)
                + ", operator="
                + str(operator)
                + ", phoneNumber="
                + str(phoneNumber)
                + ", "
                + str(networkstatus)
                + ", "
                + str(status)
            )
        else:
            final_string = (
                "status="
                + desc
                + ", deviceId='"
                + device_id
                + "', Manufacturer="
                + str(manufacturer)
                + "', model="
                + str(model)
                + ", version="
                + str(osVersion)
                + ", description="
                + str(description)
                + ", operator="
                + str(operator)
                + ", phoneNumber="
                + str(phoneNumber)
                + ",,,, "
                + str(status)
            )
        final_string = re.sub(r"^'|'$", "", final_string)
        f = open(file, "w+")
        f.write(str(final_string))
        f.close()
        sys.stdout.flush()
        return final_string
    except Exception as e:
        raise Exception("Oops!", e)
        # TODO : Dont forget to increase coma in both if else conditions if a new column is added
        if not os.path.isfile(os.path.join(TEMP_DIR, "results", device_id + ".txt")):
            if "True" in get_network_settings:
                final_string = (
                    "status=ERROR" + ",deviceId='" + device_id + "',,,,,,,,,,"
                )
            else:
                final_string = "status=ERROR" + ",deviceId='" + device_id + "',,,,,,,"
            f = open(file, "w+")
            f.write(str(final_string))
            f.close()
        return final_string


def get_list(get_dev_list):
    """get_list"""
    # Verifies each device id based on statuses
    i = 0
    split = get_dev_list.split(";")
    command = split[0]
    status = split[1]
    in_use = split[2]
    color = split[3]
    desc = split[4]
    RESPONSE = get_device_list_response(RESOURCE_TYPE, command, status, in_use)
    DEVICE_IDS = get_device_ids(RESPONSE)
    device_list = []
    if get_handset_count(RESPONSE) > 0:
        for i in range(get_handset_count(RESPONSE)):
            device_id = DEVICE_IDS[i].firstChild.data
            device_list.append(device_id + "||" + color + "||" + desc)
            device_list = [x for x in device_list if x != 0]
        if len(device_list) > 0:
            pool_size = multiprocessing.cpu_count() * 2
            pool = multiprocessing.Pool(processes=pool_size, maxtasksperchild=2)
            try:
                print(
                    "\nFound " + str(len(device_list)) + " devices with status: " + desc
                )
                sys.stdout.flush()
                output = pool.map(perform_actions, device_list)
                pool.close()
                pool.terminate()
            except Exception:
                pool.close()
                pool.terminate()
                print(traceback.format_exc())
                sys.exit(-1)


def fetch_details(i, exp_number, result, exp_list):
    """ fetches details"""
    if i == exp_number:
        if "=" in result:
            exp_list = exp_list.append(result.split("=", 1)[1].replace("'", "").strip())
        else:
            exp_list = exp_list.append("-")
    return exp_list


def fig_to_base64(fig):
    img = io.BytesIO()
    plt.savefig(img, format="png", bbox_inches="tight")
    img.seek(0)
    return base64.b64encode(img.getvalue())


def print_results(results):
    """ print_results """
    i = 0
    results.sort()
    for i in range(len(results)):
        results[i] = re.sub("Results\=$", "", results[i])
        results[i] = re.sub("[,]+", "", results[i])
        if results[i]:
            if "Available" in results[i]:
                print(colored(results[i], "green"))
            else:
                print(colored(results[i], "red"))
        i = i + 1


def validate_logo(logo):
    try:
        send_request(logo)
    except Exception as e:
        print("Exception: " + str(e))
        os.environ["company_logo"] = os.environ["perfecto_logo"]


def create_summary(df, title, column, name):
    fig = pl.figure(figsize=(15, 2))
    pl.suptitle(title)
    ax1 = pl.subplot(121, aspect="equal", facecolor="#fffffa")
    fig.patch.set_facecolor("white")
    fig.patch.set_alpha(1)
    df[column].value_counts().sort_index().plot(
        kind="pie",
        y="%",
        ax=ax1,
        autopct="%1.1f%%",
        startangle=30,
        shadow=False,
        legend=False,
        x=df[column].unique,
        fontsize=7,
    )
    pl.ylabel("")
    # plot table
    ax2 = pl.subplot(122, facecolor="#fffffa")
    ax2.patch.set_facecolor("white")
    ax2.patch.set_alpha(1)
    pl.axis("off")
    tbl = table(ax2, df[column].value_counts(), loc="center")
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(8)
    encoded = fig_to_base64(os.path.join(TEMP_DIR, "results", name + ".png"))
    summary = '<img src="data:image/png;base64, {}"'.format(encoded.decode("utf-8"))
    return summary


def create_pie(df, title, column, name):
    status = []
    for i in range(len(df[column].value_counts().sort_index().to_frame())) :
        status.append(df[column].value_counts().sort_index().to_frame().iloc[i].name)
    fig = px.pie(values=df['status'].value_counts().sort_index().to_frame().values, names=status, color=status, opacity=1, hole=0.5, width=490, height=200,  
                 color_discrete_map={'PASSED':'#35a600',
                                 'FAILED':'#f14c4c',
                                 'BLOCKED':'#4cb2ff',
                                 'UNKNOWN':'#929da5'})     
    fig.update_layout(margin=dict(t=0, b=0, l=0, r=0))
    img = fig.to_image(format="png")
    summary = '<img src="data:image/png;base64, {}"'.format(base64.b64encode(img).decode("utf-8"))
    return summary

def prepare_graph(df, column):
    """ prepare graph """
    fig = pl.figure()
    fig.patch.set_facecolor("green")
    fig.patch.set_alpha(1)
    ax = (
        df[column]
        .value_counts()
        .sort_index()
        .plot(kind="bar", fontsize=12, stacked=True, figsize=(25, 10), ylim=(0, 2))
    )
    ax.set_title(column, fontsize=20)
    ax.yaxis.set_major_formatter(ticker.FormatStrFormatter("%.2f"))
    ax.patch.set_facecolor("green")
    ax.patch.set_alpha(0.1)
    pl.yticks(df[column].value_counts(), fontsize=10, rotation=40)
    encoded = fig_to_base64(os.path.join(TEMP_DIR, "results", column + ".png"))
    return '<img src="data:image/png;base64, {}"'.format(encoded.decode("utf-8"))


"""
    Dictionary
"""


class my_dictionary(dict):
    def __init__(self):
        self = dict()

    def add(self, key, value):
        self[key] = value


"""
    Creates payload for reporting API
"""


def payloadJobAll(reportTags, oldmilliSecs, current_time_millis, jobName, jobNumber, page, boolean):
    payload = my_dictionary()
    if oldmilliSecs != 0:
        payload.add("startExecutionTime[0]", oldmilliSecs)
    if reportTags != "":
        for i, reportTaging in enumerate(reportTags.split(";")):
            payload.add("tags[" + str(i) + "]", reportTaging)
    if current_time_millis != 0:
        payload.add("endExecutionTime[0]", current_time_millis)
    payload.add("_page", page)
    if jobName != "":
        if jobName != "All Jobs":
            if jobName != "Perfecto Integrations":
                for i, job in enumerate(jobName.split(";")):
                    payload.add("jobName[" + str(i) + "]", job)
    if jobNumber != "" and boolean:
        for i, jobNumber in enumerate(jobNumber.split(";")):
            payload.add("jobNumber[" + str(i) + "]", int(jobNumber))
    print(str(payload))
    return payload


"""
    Retrieve a list of test executions within the last month
    :return: JSON object contains the executions
"""


def retrieve_tests_executions(daysOlder, page):
    current_time_millis = 0
    oldmilliSecs = 0
    global endDate
    if endDate != "":
        endTime = datetime.strptime(
            str(endDate) + " 23:59:59,999", "%Y-%m-%d %H:%M:%S,%f"
        )
        print("endExecutionTime: " + str(endTime))
        millisec = endTime.timestamp() * 1000
        current_time_millis = round(int(millisec))
    if startDate != "":
        oldmilliSecs = pastDateToMS(startDate, daysOlder)
    global reportTag
    if jobNumber != "" and jobName != "" and startDate != "" and endDate != "":
        payload = payloadJobAll(
            reportTag, oldmilliSecs, current_time_millis, jobName, jobNumber, page, False
        )
    else:
        payload = payloadJobAll(
            reportTag, oldmilliSecs, current_time_millis, jobName, jobNumber, page, True
        )
    url = "https://" + os.environ["cloudName"] + ".reporting.perfectomobile.com"
    api_url = url + "/export/api/v1/test-executions"
    # creates http geat request with the url, given parameters (payload) and header (for authentication)
    r = requests.get(
        api_url, params=payload, headers={"PERFECTO_AUTHORIZATION": os.environ["securityToken"]}
    )
    # #print(str(r.content))
    print(str(r.url))
    return r.content


def df_formatter(df):
    if len(df) < 1:
        raise Exception("Unable to find any matching executions!")
    try:
        df["startTime"] = pandas.to_datetime(df["startTime"], unit="ms")
        df["startTime"] = (
            df["startTime"].dt.tz_localize("utc").dt.tz_convert(tzlocal.get_localzone())
        )
        df["startTime"] = df["startTime"].dt.strftime("%d/%m/%Y %H:%M:%S")
    except:
        pass
    try:
        df.loc[df["endTime"] < 1, "endTime"] = int(round(time.time() * 1000))
        df["endTime"] = pandas.to_datetime(df["endTime"], unit="ms")
        df["endTime"] = (
            df["endTime"].dt.tz_localize("utc").dt.tz_convert(tzlocal.get_localzone())
        )
        df["endTime"] = df["endTime"].dt.strftime("%d/%m/%Y %H:%M:%S")
    except:
        pass
    if "month" not in df.columns:
        df["month"] = pandas.to_datetime(
            df["startTime"], format="%d/%m/%Y %H:%M:%S"
        ).dt.to_period("M")
    if "startDate" not in df.columns:
        df["startDate"] = pandas.to_datetime(
            pandas.to_datetime(df["startTime"], format="%d/%m/%Y %H:%M:%S")
            .dt.to_period("D")
            .astype(str)
        )
    if "week" not in df.columns:
        df["week"] = pandas.to_datetime(df["startDate"].dt.strftime("%Y/%m/%d")) - df[
            "startDate"
        ].dt.weekday.astype("timedelta64[D]")
    if "Duration" not in df.columns:
        df["Duration"] = pandas.to_datetime(df["endTime"]) - pandas.to_datetime(
            df["startTime"]
        )
        df["Duration"] = df["Duration"].dt.seconds
        df["Duration"] = pandas.to_datetime(df["Duration"], unit="s").dt.strftime(
            "%H:%M:%S"
        )
    if "failureReasonName" not in df.columns:
        df["failureReasonName"] = ""
    # df["name"] = '=HYPERLINK("'+df["reportURL"]+'", "'+df["name"]+'")'  # has the ability to hyperlink name in csv'
    # Filter only job and job number if dates are parameterized as well but show full histogram
    if jobNumber != "" and jobName != "":
        ori_df = df
        if ";" in jobNumber:
            df = df[df["job/number"].isin(jobNumber.split(";"))]
        else:
            df = df[df["job/number"].astype(str) == jobNumber]
    if tags != "":
        ori_df = df
        spike_cols = [col for col in df.columns if 'spike' in col]
    if startDate != "":
        name = startDate
    else:
        name = jobName + "_" + jobNumber
    df = df_to_xl(df, str(name).replace("/", "_"))
    return df


"""
    flattens the json
"""


def flatten_json(nested_json, exclude=[""]):
    """Flatten json object with nested keys into a single level.
        Args:
            nested_json: A nested json object.
            exclude: Keys to exclude from output.
        Returns:
            The flattened json object if successful, None otherwise.
    """
    out = {}

    def flatten(x, name="", exclude=exclude):
        if type(x) is dict:
            for a in x:
                if a not in exclude:
                    flatten(x[a], name + a + "/")
        elif type(x) is list:
            i = 0
            for a in x:
                flatten(a, name + str(i) + "/")
                i += 1
        else:
            out[name[:-1]] = x
    flatten(nested_json)
    return out


"""
    get final dataframe
"""


def get_final_df(files):
    df = pandas.DataFrame()
    for file in files:
        if "csv" in os.environ["xlformat"]:
            df = df.append(pandas.read_csv(file, low_memory=False))
        else:
            df = df.append(pandas.read_excel(file))
        print("Analysing file: " + str(file) + " , row count: " + str(len(df)))
    df = df_formatter(df)
    return df


"""
   gets the top failed device pass count, handset errors and device/ desktop details
"""


def getDeviceDetails(device, deviceFailCount):
    devicePassCount = 0
    errorsCount = 0
    i = 0

    for resource in resources:
        try:
            test_execution = resource  # retrieve a test execution
            # get devices which fails
            platforms = test_execution["platforms"]  # retrieve the platforms
            platform = platforms[0]
            actual_deviceID = platform["deviceId"]
            if actual_deviceID in device:
                status = test_execution["status"]
                if status in "PASSED":
                    devicePassCount += 1
                elif status in "FAILED":
                    message = test_execution["message"]
                    if "HANDSET_ERROR" in message:
                        errorsCount += 1
                deviceType = platform["deviceType"]
                if "DESKTOP" in deviceType:
                    browserInfo = platform["browserInfo"]
                    topDeviceFailureDict[device] = [
                        platform["os"] + "_" + platform["osVersion"],
                        browserInfo["browserType"]
                        + "_"
                        + browserInfo["browserVersion"],
                        devicePassCount,
                        deviceFailCount,
                        errorsCount,
                    ]
                else:
                    mobileInfo = platform["mobileInfo"]
                    topDeviceFailureDict[device] = [
                        mobileInfo["manufacturer"],
                        mobileInfo["model"],
                        devicePassCount,
                        deviceFailCount,
                        errorsCount,
                    ]
        except IndexError:
            continue
        except KeyError:
            continue
        i += 1


"""
   gets the total pass count of each failed case
"""


def getPassCount(testName):
    testNamePassCount = 0
    i = 0
    for resource in resources:
        try:
            test_execution = resource  # retrieve a test execution
            name = test_execution["name"]

            if testName in name:
                status = test_execution["status"]
                if status in "PASSED":
                    testNamePassCount += 1
        except IndexError:
            continue
        except KeyError:
            continue
        i += 1
    return testNamePassCount


"""
   gets fail and pass count of each test case and assigns it to a dict
"""


def getTCDetails(tcName, failureCount):
    topTCFailureDict[tcName] = [failureCount, getPassCount(tcName)]


"""
   calculates the percetage of a part and whole number
"""


def percentageCalculator(part, whole):
    if int(whole) > 0:
        calc = (100 * float(part) / float(whole), 0)
        calc = round(float((calc[0])), 2)
    else:
        calc = 0
    return calc


"""
   gets start date to milliseconds
"""


def pastDateToMS(startDate, daysOlder):
    dt_obj = datetime.strptime(
        startDate + " 00:00:00,00", "%Y-%m-%d %H:%M:%S,%f"
    ) - timedelta(days=daysOlder)
    millisec = dt_obj.timestamp() * 1000
    oldmilliSecs = round(int(millisec))
    return oldmilliSecs


def color_negative_red(value):
    color = "red" if value < 1 else "black"
    return "color: %s" % color


"""
   gets' Perfecto reporting API responses, creates dict for top device failures, auto suggestions and top tests failures and prepared json
"""

def prepareReport(jobName, jobNumber, reportTag):
    page = 1
    i = 0
    truncated = True
    resources = []
    resources.clear()
    print("#Parameters:")
    print("endDate: " + endDate)
    print("startDate: " + startDate)
    print("jobName: " + jobName)
    print("jobNumber: " + jobNumber)
    print("tags: " + reportTag)
    json_raw = os.environ["cloudName"] + "_API_output" +'.txt'
    open(json_raw, 'w').close
    while truncated == True:
        print(
            "Retrieving all the test executions in your lab. Current page: "
            + str(page)
            + ". Hold On!!"
        )
        executions = retrieve_tests_executions(0, page)
        # print(executions)
        # Loads JSON string into JSON object
        executions = json.loads(executions)
        if "{'userMessage': 'Failed decoding the offline token:" in str(executions):
            raise Exception("please change the offline token for your cloud")
        if "userMessage': 'Missing Perfecto-TenantId header" in str(executions):
            raise Exception("Check the cloud name and security tokens")
        if "userMessage': 'Time period is not in supported range" in str(executions):
            raise Exception("Time period is not in supported range. Check your startDate parameter")
        try:
            executionList = executions["resources"]
        except TypeError:
            print(executions)
            raise Exception(
                "Unable to find matching records for: "
                + str(criteria)
                + ", error:"
                + str(executions["userMessage"])
            )
            sys.exit(-1)
        if len(executionList) == 0:
            print("0 test executions")
            break
        else:
            # print(str(executions))
            metadata = executions["metadata"]
            truncated = metadata["truncated"]
            if page >= 1:
                resources.extend(executionList)
            else:
                resources.append(executionList)
            page += 1
    if len(resources) > 0:
        jsonDump = json.dumps(resources)
        resources = json.loads(jsonDump)
        totalTCCount = len(resources)
        with open(json_raw, "a", encoding="utf-8") as myfile:
            myfile.write(str(resources)+'\n*******************************************\n')
        print("Total executions: " + str(len(resources)))
        df = pandas.DataFrame([flatten_json(x) for x in resources])
        df = df_formatter(df)
    os.chdir(".")
    files = glob.glob("*.{}".format(os.environ["xlformat"]))
    consolidate = os.environ["consolidate"]
    if consolidate != "":
        for file in files:
            if os.path.isfile(file):
                shutil.copy2(file, consolidate)
        files = glob.iglob(os.path.join(consolidate, "*." + os.environ["xlformat"]))
    df = get_final_df(files)
    df = df.sort_values(by=["startDate"], ascending=False)
    if jobNumber != "" and jobName != "":
        if jobName != "All Jobs":
            if jobName != "Perfecto Integrations":
                ori_df = df
                df = df[df["job/name"].astype(str).isin(jobName.split(";"))]
                df = df[df["job/number"].round(0).astype(int).isin(jobNumber.split(";"))]
    if jobNumber == "" and jobName != "":
        if jobName != "All Jobs":
            if jobName != "Perfecto Integrations":
                ori_df = df
                df = df[df["job/name"].astype(str).isin(jobName.split(";"))]
    # No support for tags in consolidation
    # if reportTag != "":
    #     l = [tuple(i) for i in reportTag.split(";")]
    #     df = df[df[df.columns[pandas.Series(df.columns).str.startswith('tags/')]].apply(tuple, axis = 1).astype(str).isin(l)]
    df = df_to_xl(df, "final")
    if (len(df)) < 1:
        print("Unable to find any test executions for expected parameters")
        sys.exit(-1)

    # ggplot2 #plotly_dark #simple_white
    graphs = []
    interactive_graphs = []
    width=400
    height=400
    
    import subprocess
    if os.name == 'nt':
        DETACHED_PROCESS = 0x00000008
        subprocess.call('taskkill /F /IM Electron.exe', creationflags=DETACHED_PROCESS)
        orcaport = os.environ["orcaport"]
        subprocess.Popen(["orca", "serve", "-p", orcaport], stdout=subprocess.PIPE, shell=True)
        plotly.io.orca.config.server_url = "http://localhost:" + orcaport
        plotly.io.orca.status._props["state"] = "validated" 
    if trends == "true":
        counter = 7
        with open(live_report_filename, "a") as f:
            f.write('<div id="nestle-section">')
        duration = "weeks"
        if startDate != "":
            delta = datetime.strptime(endDate, "%Y-%m-%d") - datetime.strptime(
                startDate, "%Y-%m-%d"
            )
            if (delta.days) <= 14:
                duration = "dates"
        else:
            duration = "dates"
        joblist = []
        if "job/name" in df.columns and jobName != "":
            joblist = sorted(df["job/name"].dropna().unique())
        else:
            joblist.append("Overall!")
        for job in joblist:
            predict_df = df
            fig = []
            if job != "Overall!":
                if job in jobName or jobName == "All Jobs" or str(job).endswith("Sample"):
                    if duration == "dates":
                        fig = px.histogram(
                            df.loc[df["job/name"] == job],
                            x="startDate",
                            color="status",
                            color_discrete_map={
                                "PASSED": "limegreen",
                                "FAILED": "crimson",
                                "UNKNOWN": "#9da7f2",
                                "BLOCKED": "#e79a00",
                            },
                            hover_data=df.columns,
                            template="seaborn",
                            opacity=0.5,
                            height=height,
                            width=width
                        )
                    else:
                        fig = px.histogram(
                            df.loc[df["job/name"] == job],
                            x="week",
                            color="status",
                            hover_data=df.columns,
                            color_discrete_map={
                                "PASSED": "limegreen",
                                "FAILED": "crimson",
                                "UNKNOWN": "#9da7f2",
                                "BLOCKED": "#e79a00",
                            },
                            template="seaborn",
                            opacity=0.5,
                            height=height,
                            width=width
                        )
                    predict_df = df.loc[df["job/name"] == job]
            else:
                fig = px.histogram(
                    df,
                    x="startDate",
                    color="status",
                    color_discrete_map={
                        "PASSED": "limegreen",
                        "FAILED": "crimson",
                        "UNKNOWN": "#9da7f2",
                        "BLOCKED": "#e79a00",
                    },
                    hover_data=df.columns,
                    template="seaborn",
                    opacity=0.5,
                    height=height,
                    width=width
                )
            predict_df = (
                predict_df.groupby(["startDate"])
                .size()
                .reset_index(name="#status")
                .sort_values("#status", ascending=False)
            )
            radio = 'style="box-sizing: border-box; display: none;"'
            tabcontent = 'style="box-sizing: border-box; padding: 10px; height: auto; -moz-transition: height 1s ease; -webkit-transition: height 1s ease; -o-transition: height 1s ease; transition: height 1s ease; overflow: auto; display: inline;justify-content: center;"'
            reportDiv = (
                'style="box-sizing: border-box; overflow-x: auto; text-align: -webkit-center;"'
            )
            predictionDiv = 'style="overflow-x: auto;text-align: center;display: inline-block;float: left;background-color:white;"'
            header = 'align=center; style="box-sizing: border-box; float: left; width: 100%; padding: 1px 0; text-align: center; cursor: pointer; font-size: 16px; color: darkslategray; background-color: darkkhaki; border: 3px solid antiquewhite;"'
            if fig:
                ci_name = os.environ["ci_name"]
                tag = ""
                if ci_name != "":
                    ci_jenkins_url = os.environ["ci_jenkins_url"]
                    ci_username = os.environ["ci_username"]
                    ci_token = os.environ["ci_token"]
                    ci_href = ""
                    ci_server_url = ""
                    ci_src = ""
                    if "jenkins" in ci_name.lower(): 
                        circleCIjobs = ["MavenCircleCISample","fastlane-plugin-perfecto","FastlaneEspressoCircleCISlackSample"]
                        travisCIjobs = ["TravisSample"]
                        group = "PerfectoMobileSA"
                        if job in circleCIjobs:
                            ci_server_url = "https://circleci.com/gh/" + group
                            ci_href = ci_server_url + "/" + job
                            ci_src = ci_server_url + "/" + job + ".svg?style=shield"
                        elif job in travisCIjobs:
                            ci_server_url = "https://travis-ci.org/" + group 
                            ci_href = ci_server_url + "/" + job
                            ci_src = ci_server_url + "/" + job + ".svg?branch=master"
                        else:    
                            j = Jenkins(ci_jenkins_url, username=ci_username, password=ci_token)
                            ci_server_url = str(j).split("at ")[1]
                            ci_href = str(ci_server_url + "/job/" + job)
                            ci_src = ci_server_url + "/buildStatus/icon?job=" + job
                        print("ci_server_url: " + ci_server_url + ", ci_href: "+ ci_href + ", ci_src=" + ci_src)
                        tag = " <a href='" + ci_href + "' target='_blank'><img src='" + ci_src + "'></a>"
                fig = update_fig(fig, "histogram", job, duration)
                encoded = base64.b64encode(plotly.io.to_image(fig))
                graphs.append('<div '
                    + header
                    + '><b><center>job: '
                    + job
                    + tag
                    + '</center></b></label></div><div align="center" class="tab-content1" '
                    + tabcontent
                    + "><div " + predictionDiv + ">"
                    + '<img src="data:image/png;base64, {}"'.format(
                        encoded.decode("ascii")
                    )
                    + " alt='days or weeks summary of "
                    + job
                    + "' id='reportDiv' "
                    + reportDiv
                    + "> </img>"
                )
                interactive_graphs.append(
                        '<div style="text-align: center;"><input type="radio" id="tab'
                        + str(counter)
                        + '" name="tabs" checked=""/><label for="tab'
                        + str(counter)
                        + '">job: '
                        + job
                        + tag
                        + '</label><div class="tab-content1" style="background-color:white;">'
                        + str(fig.to_html(full_html=False, include_plotlyjs="cdn")).replace("<div",'<div style="float:left;"')    
                    )
                # interactive_graphs.append("</div>")
            if job == "Overall!" or job in jobName or jobName == "All Jobs" or str(jobName).endswith("Sample"):
                if job in jobName or jobName == "All Jobs":
                    if len(predict_df.index) > 1:
                        predict_df = predict_df.rename(
                            columns={"startDate": "ds", "#status": "y"}
                        )
                        predict_df["cap"] = int(predict_df["y"].max()) * 2
                        predict_df["floor"] = 0
                        from fbprophet import Prophet

                        with suppress_stdout_stderr():
                            m = Prophet(
                                seasonality_mode="additive",
                                growth="logistic",
                                changepoint_prior_scale=0.001,
                                weekly_seasonality=False,
                                daily_seasonality=True,
                                yearly_seasonality=False,
                            ).fit(predict_df, algorithm="Newton")
                        future = m.make_future_dataframe(periods=30)
                        future["cap"] = int(predict_df["y"].max()) * 2
                        floor = 0
                        if (int(predict_df["y"].min()) / 2) > 0:
                            floor = int(predict_df["y"].min())
                        future["floor"] = floor
                        forecast = m.predict(future)
                        forecast[["ds", "yhat", "yhat_lower", "yhat_upper"]].tail()
                        fig = plot_plotly(m, forecast, figsize=([height,width]))
                        fig = update_fig(fig, "prediction", job, duration)
                        encoded = base64.b64encode(plotly.io.to_image(fig))
                        # counter += 1
                        graphs.append('<div '
                            + predictionDiv
                            + '>'
                            + '<img src="data:image/png;base64, {}"'.format(
                                encoded.decode("ascii")
                            )
                            + " alt='prediction of "
                            + job
                            + "' id='reportDiv' "
                            + reportDiv
                            + "> </img></div><br>"
                        )
                        interactive_graphs.append(
                                '<div class="predictionDiv">'
                                + fig.to_html(full_html=False, include_plotlyjs="cdn")
                                + " </img></p></div>"
                            )
                        interactive_graphs.append("</div>")
                    else:
                        interactive_graphs.append("</div>")
                        print(
                            "Note: AI Prediction for job: "
                            + job
                            + " requires more than 2 days of data to analyze!"
                        
                        )
            counter += 1
        graphs.append("</div>")
        interactive_graphs.append("</div>")
        with open(live_report_filename, "a") as f:
            f.write("</div>")
    return graphs, interactive_graphs, df


"""
   suppress prophet logs
"""


class suppress_stdout_stderr(object):
    """
    A context manager for doing a "deep suppression" of stdout and stderr in
    Python, i.e. will suppress all print, even if the print originates in a
    compiled C/Fortran sub-function.
       This will not suppress raised exceptions, since exceptions are printed
    to stderr just before a script exits, and after the context manager has
    exited (at least, I think that is why it lets exceptions through).

    """

    def __init__(self):
        # Open a pair of null files
        self.null_fds = [os.open(os.devnull, os.O_RDWR) for x in range(2)]
        # Save the actual stdout (1) and stderr (2) file descriptors.
        self.save_fds = (os.dup(1), os.dup(2))

    def __enter__(self):
        # Assign the null pointers to stdout and stderr.
        os.dup2(self.null_fds[0], 1)
        os.dup2(self.null_fds[1], 2)

    def __exit__(self, *_):
        # Re-assign the real stdout/stderr back to (1) and (2)
        os.dup2(self.save_fds[0], 1)
        os.dup2(self.save_fds[1], 2)
        # Close the null files
        os.close(self.null_fds[0])
        os.close(self.null_fds[1])


"""
   returns a boolean if the provided string is a date or nots
"""


def is_date(string):
    try:
        parse(string)
        return True
    except ValueError:
        return False


"""
   converts datafame to excel
"""


def df_to_xl(df, filename):
    custom_columns = [
        "name",
        "status",
        "platforms/0/os",
        "platforms/0/mobileInfo/model",
        "platforms/0/browserInfo/browserType",
        "platforms/0/browserInfo/browserVersion",
        "platforms/0/osVersion",
        "failureReasonName",
        "message",
        "startTime",
        "endTime",
        "Duration",
        "job/name",
        "job/number",
        "job/branch",
        "owner",
        "reportURL",
        "platforms/0/deviceId",
        "platforms/0/deviceType",
        "platforms/0/mobileInfo/manufacturer",
        "platforms/0/screenResolution",
        "platforms/0/location",
        "platforms/0/mobileInfo/imei",
        "platforms/0/mobileInfo/phoneNumber",
        "platforms/0/mobileInfo/distributor",
        "platforms/0/mobileInfo/firmware",
        "platforms/0/selectionCriteriaV2/0/name",
        "platforms/0/selectionCriteriaV2/1/name",
        "platforms/0/selectionCriteriaV2/2/name",
        "platforms/0/selectionCriteriaV2/2/value",
        "platforms/0/customFields/0/name",
        "platforms/0/customFields/0/value",
        "tags/0",
        "tags/1",
        "tags/2",
        "tags/3",
        "tags/4",
        "tags/5",
        "tags/6",
        "tags/7",
        "tags/8",
        "tags/9",
        "tags/10",
        "tags/11",
        "tags/12",
        "tags/13",
        "tags/14",
        "tags/15",
        "tags/16",
        "tags/17",
        "tags/18",
        "tags/19",
        "tags/20",
        "id",
        "externalId",
        "uxDuration",
        "videos/0/startTime",
        "videos/0/endTime",
        "videos/0/format",
        "videos/0/streamingUrl",
        "videos/0/downloadUrl",
        "videos/0/screen/width",
        "videos/0/screen/height",
        "executionEngine/version",
        "project/name",
        "project/version",
        "automationFramework",
        "parameters/0/name",
        "parameters/0/value",
        "parameters/1/name",
        "parameters/1/value",
        "parameters/2/name",
        "parameters/2/value",
        "parameters/3/name",
        "parameters/3/value",
        "parameters/4/name",
        "parameters/4/value",
        "parameters/5/name",
        "parameters/5/value",
        "parameters/6/name",
        "parameters/6/value",
        "parameters/7/name",
        "parameters/7/value",
        "parameters/8/name",
        "parameters/8/value",
        "parameters/9/name",
        "parameters/9/value",
        "parameters/10/name",
        "parameters/10/value",
        "parameters/11/name",
        "parameters/11/value",
        "parameters/12/name",
        "parameters/12/value",
        "parameters/13/name",
        "parameters/13/value",
        "platforms/0/mobileInfo/operator",
        "platforms/0/mobileInfo/operatorCountry",
        "platforms/0/selectionCriteriaV2/3/name",
        "platforms/0/selectionCriteriaV2/3/value",
        "platforms/0/selectionCriteriaV2/4/name",
        "platforms/0/selectionCriteriaV2/4/value",
        "platforms/0/selectionCriteriaV2/5/name",
        "platforms/0/selectionCriteriaV2/5/value",
        "platforms/0/selectionCriteriaV2/6/name",
        "platforms/0/selectionCriteriaV2/6/value",
        "platforms/0/selectionCriteriaV2/7/name",
        "platforms/0/selectionCriteriaV2/7/value",
        "customFields/0/name",
        "customFields/0/value",
        "customFields/1/name",
        "customFields/1/value",
        "artifacts/0/type",
        "artifacts/0/path",
        "artifacts/0/zipped",
        "artifacts/1/type",
        "artifacts/1/path",
        "artifacts/1/contentType",
        "artifacts/1/zipped",
        "artifacts/2/type",
        "artifacts/2/path",
        "artifacts/2/zipped",
        "artifacts/0/contentType",
        "artifacts/2/contentType",
        "platforms/1/deviceId",
        "platforms/1/deviceType",
        "platforms/1/os",
        "platforms/1/osVersion",
        "platforms/1/screenResolution",
        "platforms/1/location",
        "platforms/1/mobileInfo/imei",
        "platforms/1/mobileInfo/manufacturer",
        "platforms/1/mobileInfo/model",
        "platforms/1/mobileInfo/distributor",
        "platforms/1/mobileInfo/firmware",
        "platforms/1/selectionCriteriaV2/0/name",
        "platforms/1/selectionCriteriaV2/0/value",
        "platforms/1/customFields/0/name",
        "platforms/1/customFields/0/value",
        "videos/1/startTime",
        "videos/1/endTime",
        "videos/1/format",
        "videos/1/streamingUrl",
        "videos/1/downloadUrl",
        "videos/1/screen/width",
        "videos/1/screen/height",
        "platforms/1/mobileInfo/phoneNumber",
        "month",
        "week",
        "startDate",
    ]
    df = df[df.columns.intersection(custom_columns)]
    df = df.reindex(columns=custom_columns)
    df = df.dropna(axis=1, how="all")
    filename = [filename, ".", os.environ["xlformat"]]
    if "csv" in os.environ["xlformat"]:
        df.to_csv("".join(filename), index=False)
    else:
        df.to_excel("".join(filename), index=False)
    if "csv" not in os.environ["xlformat"]:
        wb = Workbook()
        wb = load_workbook("".join(filename))
        ws = wb.worksheets[0]
        for column_cells in ws.columns:
            length = max(len(as_text(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 5
        newfilename = os.path.abspath("".join(filename))
        wb.save(newfilename)
    return df


"""
  get criteria details and values
"""


def get_report_details(item, temp, name, criteria):
    if name + "=" in item:
        temp = str(item).split("=", 1)[1]
    return str(temp), criteria


"""
  update figure
"""


def update_fig(fig, type, job, duration):
    fig.update_layout(
        title={"text": "", "y": 0.97, "x": 0.5, "xanchor": "center", "yanchor": "top"},
        xaxis_title=duration,
        yaxis_title="Test Status",
        font=dict(
            family="Trebuchet MS, Helvetica, sans-serif", size=12, color="black",
        ),
        autosize=True,
        hovermode="x unified",
        yaxis={"tickformat": ".0f"},
        xaxis_tickformat="%d/%b/%y",
    )
    fig.update_yaxes(automargin=True)
    if type == "prediction":
        fig.update_layout(
            title={"text": ""}, yaxis_title="Total tests executed",autosize=True,
        )
    return fig


""" 
get styles
"""


def get_style():
    return (
        """
            <style>

                html {{
                height:100%;
                }}
                
                .tabbed {{
                display:  flex;
                text-align: left;
                flex-wrap: wrap;
                box-shadow: 0 0 80px rgba(101, 242, 183, 0.4);
                font-size: 12px;
                font-family: "Trebuchet MS", Helvetica, sans-serif;
                margin:5px;
                width: calc(100% - 10px);
                }}
                .tabbed > input {{
                display: none;
                }}
                .tabbed > input:checked + label {{
                font-size: 14px;
                text-align: center;
                color: white;
                background-image: linear-gradient(to left, #bfee90, #bfee90, black, black,black, #bfee90, #bfee90);
                }}
                .tabbed > input:checked + label + div {{
                color:darkslateblue;
                display: block;
                }}
                .tabbed > label {{
                background-image: linear-gradient(to left, #fffeea,  #333333, #333333 ,#333333 ,#333333 , #333333, #fffeea);
                color: white;
                text-align: center;
                display: block;
                order: 1;
                flex-grow: 1;
                padding: .3%;
                }}
                .tabbed > div {{
                width: calc(100% - 10px);
                order: 2;
                flex-basis: 100%;
                display: none;
                padding: 10px;
                }}

                /* For presentation only */
                .container {{
                width: 100%;
                margin: 0 auto;
                background-color: """
        + os.environ["bgcolor"]
        + """;
                box-shadow: 0 0 20px rgba(400, 99, 228, 0.4);
                }}

                .tabbed {{
                border: 1px solid;
                }}

                hr {{
                background-color: white;
                height: 5px;
                border: 0;
                margin: 10px 0 0;
                }}
                
                hr + * {{
                margin-top: 10px;
                }}
                
                hr + hr {{
                margin: 0 0;
                }}

                .mystyle {{
                    font-size: 12pt;
                    font-family: "Trebuchet MS", Helvetica, sans-serif;
                    border-collapse: collapse;
                    border: 2px solid black;
                    margin:auto;
                    box-shadow: 0 0 80px rgba(2, 112, 0, 0.4);
                    background-color: #fffffa;
                }}

                .mystyle body {{
                font-family: "Trebuchet MS", Helvetica, sans-serif;
                    table-layout: auto;
                    position:relative;
                }}

                #slide{{
                width:100%;
                height:auto;
                }}

                #myInput, #myInput2, #myInput3 {{
                background-image: url('http://www.free-icons-download.net/images/mobile-search-icon-94430.png');
                background-position: 2px 4px;
                background-repeat: no-repeat;
                background-size: 25px 30px;
                width: 40%;
                height:auto;
                font-weight: bold;
                font-size: 12px;
                padding: 11px 20px 12px 40px;
                box-shadow: 0 0 80px rgba(2, 112, 0, 0.4);
                }}

                p {{
                text-align:center;
                color:white;
                }}

                body {{
                background-color: """
        + os.environ["bgcolor"]
        + """;
                height: 100%;
                background-repeat:  repeat-y;
                background-position: right;
                background-size:  contain;
                background-attachment: initial;
                opacity:.93;
                }}

                h4 {{
                font-family:monospace;
                }}

                @keyframes slide {{
                0% {{
                    transform:translateX(-25%);
                }}
                100% {{
                    transform:translateX(25%);
                }}
                }}

                .mystyle table {{
                    table-layout: auto;
                    width: 100%;
                    height: 100%;
                    position:relative;
                    border-collapse: collapse;
                }}

                tr:hover {{background-color:grey;}}

                .mystyle td {{
                    font-size: 12px;
                    position:relative;
                    padding: 5px;
                    width:10% !important;
                    color: black;
                    border-left: 1px solid #333;
                    border-right: 1px solid #333;
                    background: rgba(255, 253, 207, 0.58);
                    text-align: center;
                }}

                table.mystyle td:first-child {{ text-align: left; width:40% !important; }}   

                table.mystyle thead {{
                    background: grey;
                    font-size: 14px;
                    position:relative;
                    border: 1px solid black;
                }}

                table.mystyle thead th {{
                line-height: 200%;
                font-size: 14px;
                font-weight: normal;
                color: #fffffa;
                text-align: center;
                transition:transform 0.25s ease;
                }}

                table.mystyle thead th:hover {{
                    -webkit-transform:scale(1.01);
                    transform:scale(1.01);
                }}

                table.mystyle thead th:first-child {{
                border-left: none;
                }}

                .topnav {{
                overflow: hidden;
                background-color: black;
                opacity: 0.9;
                }}

                .topnav a {{
                float: right;
                display: block;
                color: #333333;
                text-align: center;
                padding: 12px 15px;
                text-decoration: none;
                font-size: 12px;
                position: relative;
                border: 1px solid #6c3;
                font-family: "Trebuchet MS", Helvetica, sans-serif;
                }}

                #summary{{
                box-shadow: 0 0 80px rgba(200, 112, 1120, 0.4);
                position: relative;
                overflow-x: scroll;
                cursor: pointer;
                padding: .1%;
                border-style: outset;
                border-radius: 1px;
                border-width: 1px;
                }}
                
                #logo{{
                box-shadow: 0 0 80px rgba(200, 112, 1120, 0.4);
                position: relative;
                cursor: pointer;
                border-style: outset;
                border-radius: 1px;
                border-width: 1px;
                }}

                .topnav a.active {{
                background-color: #333333;
                color: white;
                font-weight: lighter;
                }}

                .topnav .icon {{
                display: none;
                }}

                @media screen and (max-width: 600px) {{
                .topnav a:not(:first-child) {{display: none;}}
                .topnav a.icon {{
                    color: #DBDB40;
                    float: right;
                    display: block;
                }}
                }}

                @media screen and (max-width: 600px) {{
                .topnav.responsive {{position: relative;}}
                .topnav.responsive .icon {{
                    position: absolute;
                    right: 0;
                    top: 0;
                }}
                .topnav.responsive a {{
                    float: none;
                    display: block;
                    text-align: left;
                }}
                }}

                * {{
                box-sizing: border-box;
                }}

                img {{
                vertical-align: middle;
                }}

                .containers {{
                position: relative;
                }}

                .mySlides {{
                display:none;
                width:90%;
                }}

                #slideshow {{
                cursor: pointer;
                margin:.01% auto;
                position: relative;
                }}

                #ps{{
                height: 10%;
                margin-top: 0%;
                margin-bottom: 90%;
                background-position: center;
                background-repeat: no-repeat;
                background-blend-mode: saturation;
                }}

                #slideshow > div {{
                position: relative;
                width: 90%;
                }}

                #download {{
                background-color: #333333;
                border: none;
                color: white;
                font-size: 12px;
                cursor: pointer;
                }}

                #download:hover {{
                background-color: RoyalBlue;
                }}
                .glow {{
                    font-size: 15px;
                    color: seashell;
                    text-align: center;
                }}
                .reportDiv {{
                    overflow-x: visible;
                    text-align: -webkit-center;
                }}
                .predictionDiv {{
                    overflow-x: auto;
                    text-align: center;
                    display: inline-block;
                    float: left;
                }}
              
                #report{{
                    box-shadow: 0 0 80px rgba(145, 11, 11, 0.4);
                    overflow-x: auto;
                    min-width:70%;
                }}

                #nestle-section{{
                    float:left;
                    width:100%;
                    position:relative;
                }}

                #nestle-section label{{
                    float:left;
                    width:100%;
                    background:#333333;
                    color:rgba(245, 217, 217, 0.99);
                    padding:1px 0;
                    text-align:center;
                    cursor:pointer;
                    border:1px solid #818357;
                }}

                #nestle-section label:hover {{background-color:grey;}}

                #nestle-section .tab-content1{{
                    padding:0 10px;
                    height:0;
                    -moz-transition: height 1s ease;
                    -webkit-transition: height 1s ease;
                    -o-transition: height 1s ease;
                    transition: height 1s ease;
                    overflow:hidden;
                }}

                @media screen and (max-width: 600px) {{
                #nestle-section  input:checked + label + .tab-content1{{
                    padding: 10px;
                    height: auto;
                    -moz-transition: height 1s ease;
                    -webkit-transition: height 1s ease;
                    -o-transition: height 1s ease;
                    transition: height 1s ease;
                    overflow: auto;
                    display: block;
                }}}}
                @media screen and (min-width: 601px) {{
                #nestle-section  input:checked + label + .tab-content1{{
                    padding: 10px;
                    height: auto;
                    -moz-transition: height 1s ease;
                    -webkit-transition: height 1s ease;
                    -o-transition: height 1s ease;
                    transition: height 1s ease;
                    overflow: auto;
                    display: inline;
                    justify-content: center;  
                }}}}

                #nestle-section input:checked + label{{
                    background-color:darkkhaki;
                    color:rgb(15, 61, 16);
                    font-size:16px;
                }}#nestle-section input{{
                    display:none;
                }}
            </style>"""
    )


"""
 get failure html string 
"""

def get_failure_html_string(table):
    bg = os.environ["bgcolor"]
    string = (
        """
        <html lang="en">
       <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/4.7.0/css/font-awesome.min.css">
            <link rel="stylesheet" href="https://www.w3schools.com/w3css/4/w3.css">
    		     <head><title aria-label="Report">"""
        + str(os.environ["cloudName"]).upper() 
        + """ Failures List</title>
          <meta name="viewport" content="width=device-width, initial-scale=1">
            </head>
        <body style="background-color:"""
            + bg
            + """;">
        <style>
        .reportDiv {
                    overflow-x: visible;
                    text-align: center;
        }
        .mystyle {
            font-size: 12pt;
            font-family: "Trebuchet MS", Helvetica, sans-serif;
            border-collapse: collapse;
            border: 2px solid black;
            margin: auto;
            box-shadow: 0 0 80px rgba(2, 112, 0, 0.4);
            background-color: #fffffa;
            margin: 3%;
        }
        
        .mystyle body {
            font-family: "Trebuchet MS", Helvetica, sans-serif;
            table-layout: auto;
            position: relative;
        }
        
        .mystyle table {
            table-layout: auto;
            width: 100%;
            height: 100%;
            position: relative;
            border-collapse: collapse;
        }
        
        tr:hover {
            background-color: rgba(190, 240, 196, 0.863);
        }
        
        .mystyle td {
            box-sizing: border-box;
            font-size: 12px;
            position: relative;
            padding: 5px;
            width: 10%;
            color: black;
            border-left: 1px solid #333;
            border-right: 1px solid #333;
            background: rgba(255, 253, 207, 0.58);
        }
        
        .mystyle tr {
            text-align: center;
        }
        .mystyle thead {
            box-sizing: border-box;
            background: tan;
            color: black;
            font-size: 14px;
            position: relative;
            border: 1px solid black;
        }
        
        .mystyle th {
            box-sizing: border-box;
            line-height: 200%;
            font-size: 14px;
            background: tan;
            font-weight: bold;
            color: black;
            text-align: center;
            transition: transform 0.25s ease;
        }
        
        table.mystyle>tbody>tr>td:nth-of-type(1) {
            width: .5%;
            text-align: center;
        }
        
        table.mystyle>tbody>tr>td:nth-of-type(2) {
            width: 5%;
            text-align: left;
        }
        
        table.mystyle>tbody>tr>td:nth-of-type(3) {
            width: 10%;
            text-align: left;
        }
        table.mystyle>tbody>tr>td:nth-of-type(5) {
            width: 1%;
        }
        table.mystyle>tbody>tr>td:nth-of-type(6) {
            width: .5%;
        }
        table.mystyle>tbody>tr>td:nth-of-type(7) {
            width: .6%;
        }
    </style>
        <script src="https://ajax.googleapis.com/ajax/libs/jquery/3.4.1/jquery.min.js"></script>
            <script>
                $(document).ready(function() {
                    var table = document.getElementById("itable");
                    var rowCount = table.rows.length;
                    for (var i = 0; i < rowCount; i++) {
                        if (i >= 1) {
                            tcNameColumn = 1;
                            reportLink = 3;
                            var txt = table.rows[i].cells[tcNameColumn].innerHTML;
                            var url = table.rows[i].cells[reportLink].innerHTML;
                            var row = $('<tr></tr>')
                            var link = document.createElement("a");
                            link.href = url;
                            link.innerHTML = txt;
                            link.target = "_blank";
                            table.rows[i].cells[tcNameColumn].innerHTML = '';
                            table.rows[i].cells[tcNameColumn].appendChild(link);
                        }
                    }
                    $("table").find("tbody th").hide();
                    $("#itable").find("th, td").filter(":nth-child(" + (reportLink + 1) + ")").hide();
                    $("#search").on("keyup", debounce(function() {
                        var value = $(this).val().toLowerCase();
                        $("#itable tbody tr").filter(function() {
                            $(this).toggle($(this).text().toLowerCase().indexOf(value) > -1)
                        });
                    }, 300));

                    function debounce(func, wait, immediate) {
                        var timeout;
                        return function() {
                            var context = this,
                                args = arguments;
                            var later = function() {
                                timeout = null;
                                if (!immediate) func.apply(context, args);
                            };
                            var callNow = immediate && !timeout;
                            clearTimeout(timeout);
                            timeout = setTimeout(later, wait);
                            if (callNow) func.apply(context, args);
                        };
                    };
                });
            </script>
            <meta name="viewport" content="width=device-width, initial-scale=1">
        </head>
        <body>
        <div class="reportDiv"><br>
        <input id="search" aria-label="search" type="text" placeholder="Search/Filter..">&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</p>
        """+ table +  """ </div></body></html>""")
    return str(string)

"""
 get html string
"""


def get_html_string(graphs, tagrec, tagLinks):
    bg = os.environ["bgcolor"]
    heading = os.environ["title"]
    string = (
        """
    <html lang="en">
       <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/4.7.0/css/font-awesome.min.css">
            <link rel="stylesheet" href="https://www.w3schools.com/w3css/4/w3.css">
    		     <head><title aria-label="Report">"""
        + heading 
        + """</title>
          <meta name="viewport" content="width=device-width, initial-scale=1">
            </head>
    <body style="background-color:"""
        + bg
        + """;">
    """
        + get_style()
        + """ <div style="
          overflow-x: auto;
          text-align: center;
          color: #e5e7cc;
          background-color: #22283a;
          font-family: Helvetica, Arial, sans-serif;
          font-size: 24px;
          font-weight: bold;
          padding: 5px 5px 10px 10px;
          white-space: pre-line;box-shadow: 0 4px 8px 0 rgba(0, 0, 0, 0.2), 0 6px 20px 0 rgba(0, 0, 0, 0.19);
        "><img alt="logo" style="width:100px !important; height:30px !important;" src="https://www.perfecto.io/sites/perfecto/themes/custom/perfecto/logo.svg">
        """ + heading + """</div>"""
        + """<div style="padding:1px 0; background-color:rgba(236, 213, 171, 0.56); color:black;text-align:center;font-family:Verdana,sans-serif;font-size:16px;width:100%;cursor:pointerfont-weight:bold">
        """
        + title +
        """
        </div><div id="nestle-section">
        <input type="radio" id="tab1" name="tabs1" checked=""/><label for="tab1">Summary Report</label>
        <div class="tab-content1"> </p><div class="reportDiv"> """
        + execution_summary
        + """ alt='execution summary' id='reportDiv'> </img></br></div></div>"""
        + """<div class="reportDiv">"""
        + execution_status 
        + per_job_status
        + per_tag_status
        + """ </div></br><input type="radio" id="tab2" name="tabs" checked=""/><label for="tab2">OS Summary</label><div class="tab-content1">
          <div class="reportDiv">"""
        + monthlyStats
        + issues
        + """ <br><a href="./"""  + str(os.environ["cloudName"]) + """_failures.html" style="white-space:nowrap;text-decoration:none;background-color: rgb(195, 132, 49);color: white;padding: 8px 16px;font-family: sans-serif;border-radius: 3px;">All Failures</a>"""
        + tagLinks
        + """</p></div></div><input type="radio" id="tab4" name="tabs" checked=""/><label for="tab4">Custom Failure Reasons</label><div class="tab-content1">
          <div class="reportDiv">"""
        + failurereasons
        + """ </div></div><input type="radio" id="tab5" name="tabs" checked=""/><label for="tab5">Top Failed Tests</label><div class="tab-content1">
          <div class="reportDiv">"""
        + topfailedtable
        + """ </div>
          </div><input type="radio" id="tab6" name="tabs" checked=""/><label for="tab6">Top Recommendations</label><div class="tab-content1">
          <div class="reportDiv">"""
        + recommendations + tagrec
        + """ </div></div><div class="reportDiv">"""
        + "".join(graphs)
        + """</div></body>"""
    )
    return str(string)


"""
 get html string email
"""


def get_html_string_email(graphs, tagrec):
    bg = os.environ["bgcolor"]
    heading = os.environ["title"]
    header = 'style="box-sizing: border-box; float: left; width: 100%; padding: 1px 0; text-align: center; cursor: pointer; font-size: 16px; color: black; background-color: darkkhaki; border: 3px solid antiquewhite;"'
    tabcontent = (
        'style="box-sizing: border-box; padding: 10px; height: auto; -moz-transition: height 1s ease; -webkit-transition: height 1s ease; -o-transition: height 1s ease; transition: height 1s ease; overflow: auto; display: block;background-color:'
        + bg
        + ';"'
    )
    reportDiv = 'style="box-sizing: border-box; overflow-x: auto; text-align: center;"'

    string = (
        """
    <html lang="en">
       <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/4.7.0/css/font-awesome.min.css">
            <link rel="stylesheet" href="https://www.w3schools.com/w3css/4/w3.css">
    		     <head><title aria-label="Report">"""
        + heading
        +"""</title><meta name="viewport" content="width=device-width, initial-scale=1">
            </head>
            <style> table.mystyle td:first-child {{ text-align: left; width:40% !important; }}</style>
    <body style="box-sizing: border-box; height: 100%; background-repeat: repeat-y; background-position: right; background-size: contain; background-attachment: initial; opacity: .93; background-color:"""
        + bg
        + """;"> <div style="
          overflow-x: auto;
          text-align: center;
          color: #e5e7cc;
          background-color: #22283a;
          font-family: Helvetica, Arial, sans-serif;
          font-size: 24px;
          font-weight: bold;
          padding: 5px 5px 10px 10px;
          white-space: pre-line;box-shadow: 0 4px 8px 0 rgba(0, 0, 0, 0.2), 0 6px 20px 0 rgba(0, 0, 0, 0.19);
        "><img alt="logo" style="width:100px !important; height:30px !important;" src="https://www.perfecto.io/sites/perfecto/themes/custom/perfecto/logo.svg">
        """  + heading + """</div><div style="padding:1px 0; background-color:rgba(236, 213, 171, 0.56); color:black;text-align:center;font-family:Verdana,sans-serif;font-size:16px;width:100%;cursor:pointerfont-weight:bold">
        """
        + title +
        """<div """
        + header
        + """><b><center>Summary Report</center></b></label></div><div class="tab-content1" """
        + tabcontent
        + """>
        <div class="reportDiv" """
        + reportDiv
        + """> """
        + execution_summary
        + """ alt='execution summary' id='reportDiv' """
        + reportDiv
        + """> </img></br></div></div><div class="reportDiv" """
        + reportDiv
        + """>"""
        + execution_status
        + per_job_status_email
        + per_tag_status_email
        + """<br></div><div """
        + header
        + """</div>"""
        + """<b><center>OS Summary</center></b></label></div><div class="tab-content1" """
        + tabcontent
        + """>
        <div class="reportDiv" """
        + reportDiv
        + """>"""
        + monthlyStats
        + issues_email
        + """<br></div></div><div """
        + header
        + """><b><center>Custom Failure Reasons</center></b></label></div><div class="tab-content1" """
        + tabcontent
        + """>
        <div class="reportDiv" """
        + reportDiv
        + """>"""
        + failurereasons
        + """<br></div></div><div """
        + header
        + """><b><center>Top Failed Tests</center></b></label></div><div class="tab-content1" """
        + tabcontent
        + """>
        <div class="reportDiv" """
        + reportDiv
        + """>"""
        + topfailedtable
        + """<br></div></div><div """
        + header
        + """><b><center>Top Recommendations</center></b></label></div><div class="tab-content1" """
        + tabcontent
        + """>
        <div class="reportDiv" """
        + reportDiv
        + """>"""
        + recommendations + tagrec
        + """ </div></div></div><br><div class="reportDiv">"""
        + "".join(graphs)
        + """</div><br></body>"""
    )
    return str(string)


def prepare_html(user_html, table3, day):
    """ prepare_html """
    print(colored("\nFinal Devices list:", "magenta"))
    # copies all device status to final summary
    for r, d, f in os.walk(os.path.join(TEMP_DIR, "results")):
        for file in f:
            if ".txt" in file:
                with open(os.path.join(r, file)) as f:
                    with open(os.path.join(r, "Final_Summary.txt"), "a") as f1:
                        for line in f:
                            f1.write(line)
                            f1.write("\n")
    file = os.path.join(TEMP_DIR, "results", "Final_Summary.txt")
    try:
        f = open(file, "r")
    except FileNotFoundError:
        raise Exception("No devices found/ Re-check your arguments")
        sys.exit(-1)
    result = f.read()
    f.close()
    print_results(result.split("\n"))
    if "true" in os.environ["PREPARE_ACTIONS_HTML"]:
        results = result.split("\n")
        # TODO Add a new list and number below & a new_dict if a new column is added
        new_dict = {}
        deviceids = []
        status = []
        description = []
        manufacturer = []
        model = []
        osVersion = []
        operator = []
        phonenumber = []
        airplanemode = []
        wifi = []
        data = []
        action_results = []
        for result in results:
            if len(result) > 0:
                new_result = result.split(",")
                new_list = []
                i = 0
                for result in new_result:
                    fetch_details(i, 0, result, status)
                    fetch_details(i, 1, result, deviceids)
                    fetch_details(i, 2, result, manufacturer)
                    fetch_details(i, 3, result, model)
                    fetch_details(i, 4, result, osVersion)
                    fetch_details(i, 5, result, description)
                    fetch_details(i, 6, result, operator)
                    fetch_details(i, 7, result, phonenumber)
                    fetch_details(i, 8, result, airplanemode)
                    fetch_details(i, 9, result, wifi)
                    fetch_details(i, 10, result, data)
                    fetch_details(i, 11, result, action_results)
                    new_list.append(result)
                    i = i + 1
        pandas.set_option("display.max_columns", None)
        pandas.set_option("display.max_colwidth", 100)
        pandas.set_option("colheader_justify", "center")
        get_network_settings = os.environ["GET_NETWORK_SETTINGS"]
        reboot = os.environ["REBOOT"]
        cleanup = os.environ["CLEANUP"]
        if "True" in get_network_settings or "True" in reboot or "True" in cleanup:
            new_dict = {
                "Status": status,
                "Device Id": deviceids,
                "Manufacturer": manufacturer,
                "Model": model,
                "OS Version": osVersion,
                "Description": description,
                "Operator": operator,
                "Phone number": phonenumber,
                "AirplaneMode": airplanemode,
                "Wifi": wifi,
                "Data": data,
                "Results": action_results,
            }
        else:
            new_dict = {
                "Status": status,
                "Device Id": deviceids,
                "Manufacturer": manufacturer,
                "Model": model,
                "OS Version": osVersion,
                "Description": description,
                "Operator": operator,
                "Phone number": phonenumber,
            }
        df = pandas.DataFrame(new_dict)
        df = df.sort_values(by="Manufacturer")
        df = df.sort_values(by="Model")
        df = df.sort_values(by="Status")
        df.reset_index(drop=True, inplace=True)
        device_list_parameters = os.environ["DEVICE_LIST_PARAMETERS"]
        cloudname = os.environ["cloudName"]
        current_time = datetime.now().strftime("%c")
        title = (
            cloudname.upper()
            + " cloud status summary of "
            + device_list_parameters
            + " @ "
            + current_time
        )
        summary = create_summary(df, title, "Status", "device_summary")
        plt.close("all")

        df = df.sort_values(by="Model")
        df = df.sort_values(by="Status")
        # skipping csv output as we now have full device list api response
        #         df.to_csv(os.path.join(TEMP_DIR , 'results','output.csv'), index=False)
        # Futuristic:
        #     le = preprocessing.LabelEncoder()
        #     #convert the categorical columns into numeric
        #     dfs = df.copy()
        #     encoded_value = le.fit_transform(dfs['Device Id'])
        #     dfs['Device Id'] = le.fit_transform(dfs['Device Id'])
        #     dfs['Status'] = le.fit_transform(dfs['Status'])
        #     dfs['Model'] = le.fit_transform(dfs['Model'])
        #     dfs['OS Version'] = le.fit_transform(dfs['OS Version'])
        #     dfs['Operator'] = le.fit_transform(dfs['Operator'])
        #     dfs['Phone number'] = le.fit_transform(dfs['Phone number'])
        #     if  "True" in get_network_settings or  "True" in reboot or  "True" in cleanup:
        #         dfs['AirplaneMode'] = le.fit_transform(dfs['AirplaneMode'])
        #         dfs['Wifi'] = le.fit_transform(dfs['Wifi'])
        #         dfs['Data'] = le.fit_transform(dfs['Data'])
        #         dfs['Results'] = le.fit_transform(dfs['Results'])
        #     print(dfs)
        #     cols = [col for col in dfs.columns if col not in ['Status','Phone number', 'OS Version', 'Model', 'Operator']]
        #     data = dfs[cols]
        #     target = dfs['Status']
        #     print(data)
        #     print(target)

        html_string = (
            """
        <html lang="en">
          <head>
    	  <meta name="viewport" content="width=device-width, initial-scale=1">
           <meta content="text/html; charset=iso-8859-2" http-equiv="Content-Type">
    		<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/4.7.0/css/font-awesome.min.css">
            <link rel="stylesheet" href="https://www.w3schools.com/w3css/4/w3.css">
    		     <head><title>"""
            + cloudname.upper()
            + """ Cloud Status</title>
          <script src="https://ajax.googleapis.com/ajax/libs/jquery/3.4.1/jquery.min.js"></script>
            <script>
              $(document).ready(function(){{
                document.getElementById("tabbed-device").click();
            }});
            $(document).ready(function(){{
                // Add smooth scrolling to all links
                $("a").on('click', function(event) {{

                    // Make sure this.hash has a value before overriding default behavior
                    if (this.hash !== "") {{
                    // Prevent default anchor click behavior
                    event.preventDefault();

                    // Store hash
                    var hash = this.hash;

                    // Using jQuery's animate() method to add smooth page scroll
                    // The optional number (800) specifies the number of milliseconds it takes to scroll to the specified area
                    $('html, body').animate({{
                        scrollTop: $(hash).offset().top
                    }}, 800, function(){{
                
                        // Add hash (#) to URL when done scrolling (default click behavior)
                        window.location.hash = hash;
                    }});
                    }} // End if
                }});
            }});
            $(document).ready(function(){{
              $("#myInput").on("keyup", function() {{
                var value = $(this).val().toLowerCase();
                $("#devicetable tbody tr").filter(function() {{
                  $(this).toggle($(this).text().toLowerCase().indexOf(value) > -1)
                }});
              }});
            }});
                      $(document).ready(function(){{
              $("#myInput2").on("keyup", function() {{
                var value = $(this).val().toLowerCase();
                $("#usertable tbody tr").filter(function() {{
                  $(this).toggle($(this).text().toLowerCase().indexOf(value) > -1)
                }});
              }});
            }});
            $(document).ready(function(){{
              $("#myInput3").on("keyup", function() {{
                var value = $(this).val().toLowerCase();
                $("#repotable tbody tr").filter(function() {{
                  $(this).toggle($(this).text().toLowerCase().indexOf(value) > -1)
                }});
              }});
            }});
            </script>
    		<script type="text/javascript">
    	           $(document).ready(function(){{
                   $("#slideshow > div:gt(0)").show();
    				$("tbody tr:contains('Disconnected')").css('background-color','#fcc');
    				$("tbody tr:contains('ERROR')").css('background-color','#fcc');
    				$("tbody tr:contains('Un-available')").css('background-color','#fcc');
    				$("tbody tr:contains('Busy')").css('background-color','#fcc');
                    var table = document.getElementById("devicetable");
    				var rowCount = table.rows.length;
    				for (var i = 0; i < rowCount; i++) {{
    					if ( i >=1){{
                        available_column_number = 0;
                        device_id_column_number = 1;
    						if (table.rows[i].cells[available_column_number].innerHTML == "Available") {{
                                for(j = 0; j < table.rows[0].cells.length; j++) {{
    								table.rows[i].cells[j].style.backgroundColor = '#e6fff0';
                                        if (table.rows[i].cells[(table.rows[0].cells.length - 1)].innerHTML.indexOf("failed") > -1) {{
                                                table.rows[i].cells[j].style.color = '#660001';
                                                table.rows[i].cells[j].style.backgroundColor = '#FFC2B5';
                                        }}
                                 }}
    							var txt = table.rows[i].cells[device_id_column_number].innerHTML;
    							var url = 'https://"""
            + cloudname.upper()
            + """.perfectomobile.com/nexperience/main.jsp?applicationName=Interactive&id=' + txt;
    							var row = $('<tr></tr>')
    							var link = document.createElement("a");
    							link.href = url;
    							link.innerHTML = txt;
    							link.target = "_blank";
    							table.rows[i].cells[device_id_column_number].innerHTML = "";
    							table.rows[i].cells[device_id_column_number].appendChild(link);
    						}}else{{
    							for(j = 0; j < table.rows[0].cells.length; j++) {{
    								table.rows[i].cells[j].style.color = '#660001';
                                         table.rows[i].cells[j].style.backgroundColor = '#FFC2B5';
    							}}
    						}}
    					}}
    				}}
                 }});
                 function myFunction() {{
                  var x = document.getElementById("myTopnav");
                  if (x.className === "topnav") {{
                    x.className += " responsive";
                  }} else {{
                    x.className = "topnav";
                  }}
                }}
                function zoom(element) {{
				         var data = element.getAttribute("src");
						 let w = window.open('about:blank');
						 let image = new Image();
						 image.src = data;
						 setTimeout(function(){{
						   w.document.write(image.outerHTML);
						 }}, 0);
				     }}
                function autoselect(element) {{
                     var data = element.getAttribute("id");
                     document.getElementById(data + "-1").checked = true;
                }}     
    		</script>

    		<meta name="viewport" content="width=device-width, initial-scale=1">
            </head>
            """
            + get_style()
            + """
          <body bgcolor="#FFFFED">
    	  	<div class="topnav" id="myTopnav">
    		  <a href="result.html" class="active">Home</a>
    		  <a href="https://"""
            + cloudname.upper()
            + """.perfectomobile.com" target="_blank" class="active">"""
            + cloudname.upper()
            + """ Cloud</a>
              <a href="https://developers.perfectomobile.com" target="_blank" class="active">Docs</a>
              <a href="https://www.perfecto.io/services/professional-services-implementation" target="_blank" class="active">Professional Services</a>
    		  <a href="https://support.perfecto.io/" target="_blank" class="active">Perfecto Support</a>
    		  <a href="javascript:void(0);" aria-label="first link" class="icon" onclick="myFunction()">
    			<i class="fa fa-bars"></i>
    		  </a>
    		</div>

            <div style="text-align: center">
                
                <div class="container">
                    <div class="tabbed">
                        <input type="radio" id="tabbed-tab-1-1" onClick='autoselect(this)' name="tabbed-tab-1" checked><label for="tabbed-tab-1-1">Users</label>
                        <div>
                            <div class="tabbed">
                                <input type="radio" id="tabbed-tab-1-1-1" name="tabbed-tab-1-1" checked><label for="tabbed-tab-1-1-1">List</label>
                                <div align="center">
                                
                                <a href="https://"""
            + cloudname.upper()
            + """.perfectomobile.com" target="_blank" class="site-logo">
                                <img id="logo" src="""
            + os.environ["company_logo"]
            + """ style="margin:1%;" alt="Company logo" ></a> 
            <div class="reportDiv">
                                """
            + create_summary(user_html, "Users list Status", "status", "user_summary")
            + """ alt='user_summary' id='summary' onClick='zoom(this)'></img></br></p></div>
                                <input id="myInput2" aria-label="search" type="text" placeholder="Search..">&nbsp;&nbsp;&nbsp;
                                <a id ="download" href="./get_users_list.xlsx" aria-label="A link to users .xlsx file is present." class="btn"><i class="fa fa-download"></i> Users List</a>
                                </br> </br>
                                <div style="overflow-x:auto;">
                                    {table2}
                                </div>
                                
                                </div>
                             </div>
                        </div>
                    <input type="radio" id="tabbed-tab-1-2" onClick='autoselect(this)' name="tabbed-tab-1" checked><label id="tabbed-device" for="tabbed-tab-1-2">Device</label>
                    <div>
                        <div class="tabbed">
                            <input type="radio" id="tabbed-tab-1-2-1" name="tabbed-tab-1-1" checked><label for="tabbed-tab-1-2-1">List</label>
                            <div  align="center">
                            
                                <a href="https://"""
            + cloudname.upper()
            + """.perfectomobile.com" target="_blank" class="site-logo">
                                <img id="logo" src="""
            + os.environ["company_logo"]
            + """ style="margin:1%;" alt="Company logo" ></a> 
                                <div class="reportDiv">"""
            + summary
            + """ alt='summary' id='summary' onClick='zoom(this)'></img> </br></p></div>
                                <input id="myInput" aria-label="search" type="text" placeholder="Search..">&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
                                    <a id ="download" href="./get_devices_list.xlsx" aria-label="A link to a .xlsx file is present." class="btn"><i class="fa fa-download"></i> Full Devices List</a>
                                    </br> </br>
                                        <div style="overflow-x:auto;">
                                            {table}
                                        </div>
                                    </br>
                                    
                        </div>
                        <input type="radio" id="tabbed-tab-1-2-2" name="tabbed-tab-1-1"><label for="tabbed-tab-1-2-2">Graphs</label>
                            <div align="center">

                                  <div style="overflow-x:auto;height:90%">
                                    <div class="containers" align="center" id = "slideshow">
                                        <div class="w3-content w3-section"  style="max-width:90%; max-height:90%;height:90%;width:90%;">
                                        """
            + prepare_graph(df, "Manufacturer")
            + """ alt="Device Status" class="mySlides"  onClick='zoom(this)' id="slide">
                                        """
            + prepare_graph(df, "Model")
            + """ alt="Model" class="mySlides"  onClick='zoom(this)' id="slide">
                                        """
            + prepare_graph(df, "OS Version")
            + """ alt="Version" class="mySlides" onClick='zoom(this)' id="slide">
                                        """
            + prepare_graph(df, "Operator")
            + """ alt="Operator" class="mySlides" onClick='zoom(this)' id="slide">
                                        """
            + prepare_graph(df, "Description")
            + """ alt="Description" class="mySlides"  onClick='zoom(this)' id="slide">
                                        </div>
                                    </div>
                                </div>
                                    
                            </div>
                    </div>
                </div>
                        {table3}
                    </div>
                    <a target="_blank" style="font-size:12;font-family:"Trebuchet MS", Helvetica, sans-serif;color:powderblue;" href="https://clearbit.com">Logos provided by Clearbit</a>
                </div>
              <script>
             
              var myIndex = 0;
              carousel();

              function carousel() {{
                var i;
                var x = document.getElementsByClassName("mySlides");
                for (i = 0; i < x.length; i++) {{
                  x[i].style.display = "none";
                }}
                myIndex++;
                if (myIndex > x.length) {{myIndex = 1}}
                x[myIndex-1].style.display = "block";
                setTimeout(carousel, 2000); // Change image every 2 seconds
              }}
              </script>
          </body>
        </html>
        """
        )

        # OUTPUT AN HTML FILE
        clean_repo = os.environ["clean_repo"]
        with open(os.path.join(TEMP_DIR, "output", "result.html"), "w") as f:
            if "NA" != clean_repo:
                heading = (
                    """<input type="radio" onClick='autoselect(this)' id="tabbed-tab-1-3" name="tabbed-tab-1"><label for="tabbed-tab-1-3">Repository</label>
                            <div>
                            <div class="tabbed">
                                <input type="radio" id="tabbed-tab-1-3-1" name="tabbed-tab-1-1" checked><label for="tabbed-tab-1-3-1">List</label>
                                <div align="center">
                            <b><h4><p style="color:white">"""
                    + cloudname.upper()
                    + """ Media Repository Cleanup Status for files older than """
                    + str(day)
                    + """ days: Total - """
                    + str(table3.shape[0])
                    + """</p></font></h4>
                             <input id="myInput3" aria-label="search" type="text" placeholder="Search.."></br> </br>
                            <div style="overflow-x:auto;"></b>"""
                )
                f.write(
                    html_string.format(
                        table=df.to_html(
                            classes="mystyle", table_id="devicetable", index=False
                        ),
                        table2=user_html.to_html(
                            classes="mystyle",
                            table_id="usertable",
                            justify="justify-all",
                            index=False,
                        ),
                        table3=heading
                        + table3.to_html(
                            classes="mystyle", table_id="repotable", index=False
                        )
                        + "</div></div></div></div></br>",
                    )
                )
            else:
                f.write(
                    html_string.format(
                        table=df.to_html(
                            classes="mystyle", table_id="devicetable", index=False
                        ),
                        table2=user_html.to_html(
                            classes="mystyle",
                            table_id="usertable",
                            justify="justify-all",
                            index=False,
                        ),
                        table3="",
                    )
                )
        webbrowser.open(
            "file://" + os.path.join(TEMP_DIR, "output", "result.html"), new=0
        )
        print("Results: file://" + os.path.join(TEMP_DIR, "output", "result.html"))


def send_request_repo(url, content):
    try:
        response = urllib.request.urlopen(url.replace(" ", "%20"))
    except Exception as e:
        return e
    sys.stdout.flush()
    return response


def send_request_for_repository(url, content, key):
    response = send_request_repo(url, content)

    if "500" in str(response):
        print(url)
        raise RuntimeError(
            "Failed to list repository items - Repository item: "
            + key
            + "  was not found in media repository, url: "
            + str(url)
        )
        sys.exit(-1)
    text = response.read().decode("utf-8")
    map = json.loads(text)
    return map


def getActualDate(map):
    # date is fetched here
    try:
        date = map["item"]["creationTime"]["formatted"]
    except KeyError:
        return ""
    dateOnly = date.split("T")
    return datetime.strptime(dateOnly[0], "%Y-%m-%d")


def getPastDate(days):
    # Logic for fetching past days based on user preference
    today = datetime.today()
    pastDate = timedelta(days=int(days))
    return today - pastDate


def sendAPI(resource_type, resource_key, operation):
    url = get_url(str(resource_type), resource_key, operation)
    # admin = os.environ["repo_admin"]
    # if "true" in admin.lower():
    #     url += "&admin=" + "true"
    return send_request_for_repository(url, "", resource_key)

def sendAPI_repo(resource_type, resource_key, operation):
    url = get_url(str(resource_type), "", operation)
    print("\nRepository API Raw url: \n" + url)
    return send_request_for_repository(url, "", resource_key)


def fetch_details_repo(i, exp_number, result, exp_list):
    """ fetches details"""
    if i == exp_number:
        if ":" in result:
            exp_list = exp_list.append(result.split(":", 1)[1].replace("'", "").strip())
        else:
            exp_list = exp_list.append("-")
    return exp_list


def run_commands(value):
    # Get date of repository items
    FINAL_LIST = []
    DAYS = os.environ["repo_days"]
    DELETE = os.environ["repo_delete"]
    map = sendAPI(os.environ["repo_resource_type"], value, "info")
    print(str(map))
    actualDate = getActualDate(map)
    if not (str(actualDate) == ""):
        expectedDate = getPastDate(DAYS)
        expDate = str.split(str(expectedDate), " ")
        actDate = str(str.split(str(actualDate), " ")[0])
        # DELETES the item if older than expected date
        if actualDate < expectedDate:
            print(
                colored(
                    "File: "
                    + value
                    + " with actual creation date: "
                    + actDate
                    + " was created before "
                    + str(DAYS)
                    + " days.",
                    "red",
                )
            )
            # DELETE item from the repository
            if DELETE.lower() == "true":
                map = sendAPI(os.environ["repo_resource_type"], value, "delete")
                status = map["status"]
                if status != "Success":
                    FINAL_LIST.append(
                        "File:"
                        + value
                        + ";Created on:"
                        + actDate
                        + ";Comparison:is older than;Days:"
                        + DAYS
                        + ";Deleted?:Unable to delete!;"
                    )
                    raise RuntimeError("Repository item " + value + " was not deleted")
                    sys.exit(-1)
                else:
                    FINAL_LIST.append(
                        "File:"
                        + value
                        + ";Created on:"
                        + actDate
                        + ";Comparison:is older than;Days:"
                        + DAYS
                        + ";Deleted?:Yes;"
                    )
            else:
                FINAL_LIST.append(
                    "File:"
                    + value
                    + ";Created on:"
                    + actDate
                    + ";Comparison:is older than;Days:"
                    + DAYS
                    + ";Deleted?:No;"
                )
        else:
            print(
                colored(
                    "File: "
                    + value
                    + " with actual creation date: "
                    + actDate
                    + " was created within the last "
                    + str(DAYS)
                    + " days.",
                    "green",
                )
            )
    #                   FINAL_LIST.append('File:' + value + ';Created on:' + actDate + ';Comparison:is younger than;Days:' + DAYS + ';Deleted?:No;')
    fileName = uuid.uuid4().hex + ".txt"
    file = os.path.join(TEMP_DIR, "repo_results", fileName)
    f = open(file, "w+")
    f.write(str(FINAL_LIST))
    f.close()


def manage_repo(resource_key):
    # Get list of repository items
    map = sendAPI_repo(os.environ["repo_resource_type"], resource_key, "list")
    try:
        itemList = map["items"]
        itemList = [x for x in itemList if x.startswith(resource_key)] 
        sys.stdout.flush()
        print("Item list: " + str(itemList))
    except:
        raise RuntimeError(
            "There are no List of repository items starting with the folder names: " + resource_key
        )
        sys.exit(-1)
 
    # debug
    #     for value in itemList:
    #         run_commands(value)
    pool_size = multiprocessing.cpu_count() * 2
    repo_folder_pool = multiprocessing.Pool(processes=pool_size, maxtasksperchild=2)
    try:
        FINAL_LIST = repo_folder_pool.map(run_commands, itemList)
        repo_folder_pool.close()
        repo_folder_pool.terminate()
    except Exception:
        repo_folder_pool.close()
        repo_folder_pool.terminate()
        print(traceback.format_exc())
        sys.exit(-1)


def deleteOlderFiles(resource_type, delete, admin, repo_paths, days):
    os.environ["repo_delete"] = delete
    os.environ["repo_days"] = days
    os.environ["repo_resource_type"] = resource_type
    os.environ["repo_admin"] = admin
    create_dir(os.path.join(TEMP_DIR, "repo_results"), True)
    I = 0
    REPO_LIST = repo_paths.split(",")
    # debug:
    #     for repo in REPO_LIST:
    #         manage_repo(repo)
    procs = []
    for li in REPO_LIST:
        proc = Process(target=manage_repo, args=(str(li),))
        procs.append(proc)
        proc.start()
    try:
        for proc in procs:
            proc.join()
        for proc in procs:
            proc.terminate()
    except Exception:
        proc.terminate()
        print(traceback.format_exc())
        sys.exit(-1)

    for r, d, f in os.walk(os.path.join(TEMP_DIR, "repo_results")):
        for file in f:
            if ".txt" in file:
                with open(os.path.join(r, file)) as f:
                    with open(os.path.join(r, "Final_Repo.txt"), "a") as f1:
                        for line in f:
                            f1.write(line)
                            f1.write("\n")
    file = os.path.join(TEMP_DIR, "repo_results", "Final_Repo.txt")
    try:
        f = open(file, "r")
        result = f.read()
        f.close()
    except FileNotFoundError:
        result = ""
        pass
    
    FINAL_LIST = result.split("\n")

    file = []
    created = []
    comparison = []
    days = []
    deleted = []
    final_dict = {}
    for lists in FINAL_LIST:
        if lists is not None:
            if len(lists) > 0:
                new_result = str(lists).split(";")
                i = 0
                for result in new_result:
                    if "Deleted?:" in result:
                        fetch_details_repo(
                            i,
                            new_result.index(result, i),
                            str(result).replace("]", ""),
                            deleted,
                        )
                    if "File:" in result:
                        fetch_details_repo(i, new_result.index(result, i), result, file)
                    if "Created on:" in result:
                        fetch_details_repo(
                            i, new_result.index(result, i), result, created
                        )
                    if "Comparison:" in result:
                        fetch_details_repo(
                            i, new_result.index(result, i), result, comparison
                        )
                    if "Days:" in result:
                        fetch_details_repo(i, new_result.index(result, i), result, days)
                    i = i + 1
    pandas.set_option("display.max_columns", None)
    pandas.set_option("display.max_colwidth", 100)
    pandas.set_option("colheader_justify", "center")
    final_dict = {
        "File": file,
        "Created On": created,
        "Comparison": comparison,
        "Days": days,
        "Deleted?": deleted,
    }
    df = pandas.DataFrame(final_dict)
    df = df.sort_values(by="File")
    df.style.set_properties(**{"text-align": "left"})
    sys.stdout.flush()
    return df


def create_dir(directory, delete):
    """
    create Dir
    """
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
        else:
            if delete:
                shutil.rmtree(directory)
                os.makedirs(directory)
    except Exception as e:
        print(colored(e, "red"))
        sys.exit(-1)


def style_df_email(df):
    return (
        df.replace(
            '<table border="1" class="dataframe mystyle" id="report">',
            '<table border="1" class="dataframe mystyle" id="report" style="box-sizing: border-box; font-size: 12pt; font-family:Trebuchet MS, Helvetica, sans-serif; border-collapse: collapse; border: 2px solid black;  margin-left:5%;margin-right:5%; margin:auto;background-color: #fffffa; box-shadow: 0 0 30px rgba(145, 11, 11, 0.4); overflow-x: auto; min-width: 60%; max-width:80%;" bgcolor="#fffffa">',
        )
        .replace("<tr>", '<tr style="box-sizing: border-box;" align="center">')
        .replace(
            "<thead>",
            '<thead style="box-sizing: border-box; background: tan; color:black;font-size: 14px; position: relative; border: 1px solid black;">',
        )
        .replace(
            "<th>",
            '<th style="box-sizing: border-box; line-height: 200%; font-size: 14px; background: tan; font-weight: bold; color: black; text-align: center; transition: transform 0.25s ease;" align="center">',
        )
        .replace("<tbody>", '<tbody style="box-sizing: border-box;">')
        .replace(
            "<td>",
            '<td style="box-sizing: border-box; font-size: 12px; position: relative; padding: 5px; color: black; border-left: 1px solid #333; border-right: 1px solid #333; background: rgba(255, 253, 207, 0.58);width:15%;" align="center">',
        )
    )

def createHref(jobName, tag, status, content):
    jobList = ""
    tagList = ""
    jobNumList = ""
    if(isinstance(content, int) and int(content) == 0):
        return content
    else:
        if(startDate):
            startExecutionTime =  pastDateToMS(startDate, 0)
            endExecutionTime = round(int(datetime.strptime(
                str(endDate) + " 23:59:59,999", "%Y-%m-%d %H:%M:%S,%f"
            ).timestamp() * 1000))
            url = "https://"+ os.environ["cloudName"] + ".app.perfectomobile.com/reporting/library?startExecutionTime[0]=" + str(startExecutionTime) + "&startExecutionTime[1]=custom&endExecutionTime[0]=" + str(endExecutionTime)
        else:
            url = "https://"+ os.environ["cloudName"] + ".app.perfectomobile.com/reporting/library?"
        if(jobName):
            for i, job in enumerate(jobName.split(";")):
                    jobList = jobList + "&jobName[" + str(i) + "]=" + job
            url = url + jobList
        if(jobNumber):
            for i, jobNum in enumerate(jobNumber.split(";")):
                jobNumList = jobNumList + "&jobNumber[" + str(i) + "]=" + jobNum
            url = url + jobNumList
        if(tag):
            for i, job in enumerate(tag.split(";")):
                    tagList = tagList + "&tags[" + str(i) + "]=" + job
            url = url + tagList
        if(status):
            url = url + '&status[0]='+status+'&_search=&_searchViewAll=false'
        else:
            url = url + '&_search=&_searchViewAll=false'
            
        url = '<a target="_blank" href="' + str(url) + '">' + str(content) + "</a>"
        return url

def get_recommendations(df, failed_blocked, topfailedTCNames, failed, passed, blocked, name, bool):
    labIssuesCount = 0
    totalFailCount = 0
    totalPassCount = 0
    totalUnknownCount = 0
    totalTCCount = 0
    scriptingIssuesCount = 0
    appCrashIssuesCount = 0
    testDataIssuesCount = 0
    environmentIssuesCount = 0
    orchestrationIssuesCount = 0
    cleanedFailureList = {}
    # recommendations
    totalFailCount = failed.shape[0]
    totalPassCount = passed.shape[0]
    blockedCount = blocked.shape[0]
    # failures count
    failuresmessage = []
    failureListFileName = name + "_failures" +'.txt'
    print("transfering all failure reasons to: %s" % (os.path.join(os.path.abspath(os.curdir), failureListFileName)))
    open(failureListFileName, 'w').close
    totalUnknownCount = df[(df["Test Status"] == "UNKNOWN")].shape[0]
    totalTCCount = df.shape[0]
    if len(failed_blocked) > 0:
        failuresmessage = (
            failed_blocked.groupby(["message"])
            .size()
            .reset_index(name="#Failed")
            .sort_values("#Failed", ascending=False)
        )
            #Get all errors and links
        i = 1
        allFailuresResultLink = []
        for error, count in failuresmessage.itertuples(index=False):
            report_link = df.loc[df['message'].str.startswith(str(error.strip()) , na=False), "reportURL"].iloc[0]
            testName = df.loc[df['reportURL'].str.startswith(str(report_link) , na=False), "name"].iloc[0]
            job = df.loc[df['reportURL'].str.startswith(str(report_link) , na=False), "job/name"].iloc[0]
            jobNumber = df.loc[df['reportURL'].str.startswith(str(report_link) , na=False), "job/number"].iloc[0]
            owner = df.loc[df['reportURL'].str.startswith(str(report_link) , na=False), "owner"].iloc[0]
            miniJson = {
                "#": i,
                "Test Name": testName,
                "Error": error,
                "link": report_link,
                "Job Name": str(job),
                "Job Number": str(jobNumber),
                "Owner": str(owner),
            }
            allFailuresResultLink.append(miniJson)
            i+=1
        allFailuresResultLink = json.dumps(allFailuresResultLink)
        output = json2html.convert(json = allFailuresResultLink)
        file = open(name + '_failures.html', 'w') 
        file.write(get_failure_html_string(str(output).replace('<table border="1">','<table class="dataframe mystyle" border="1" id="itable">')))
        file.close() 

        global labIssues
        global orchestrationIssues
        for commonError, commonErrorCount in failuresmessage.itertuples(
            index=False
        ):
            for labIssue in labIssues:
                if re.search(labIssue, commonError):
                    labIssuesCount += commonErrorCount
                    break
            for orchestrationIssue in orchestrationIssues:
                if re.search(orchestrationIssue, commonError):
                    orchestrationIssuesCount += commonErrorCount
                    break
            error = str(commonError)
            regex = ""
            if os.environ["regex"] != "":
                regex = "|" + os.environ["regex"]
            regEx_Filter = "Build info:|For documentation on this error|at org.xframium.page|Scenario Steps:| at WebDriverError|\(Session info:|XCTestOutputBarrier\d+|\s\tat [A-Za-z]+.[A-Za-z]+.|View Hierarchy:|Got: |Stack Trace:|Report Link|at dalvik.system|Output:\nUsage|t.*Requesting snapshot of accessibility|\{ Error\:|at\sendReadableNT|at\sFunction|\sat\smakeRequest|at\sObject\.\_errnoException|\"stack\"\:|('|)\n.*Error\:\s|at\sRequest.callback|\n\s+at\s" + regex
            if re.search(regEx_Filter, error, re.MULTILINE):
                if re.search("('|)(\n|)AssertionError\:.*\n\s+at.*\n.*at", error, re.MULTILINE):
                    error = "at".join(error.split("at", 2)[:2]).replace(r'\n','\n')

                else:
                    error = str(re.compile(regEx_Filter).split(error)[0])
                    if "An error occurred. Stack Trace:" in error:
                        error = error.split("An error occurred. Stack Trace:")[1]
            if re.search("error: \-\[|Fatal error:", error):
                error = str(
                    re.compile("error: \-\[|Fatal error:").split(error)[1]
                )
            if error.strip() in cleanedFailureList:
                cleanedFailureList[error.strip()] += 1
            else:
                cleanedFailureList[error.strip()] = commonErrorCount
            appCrashIssuesCount =  len(df.loc[df['Custom Failure Reason'] == "Application crashed"])
            environmentIssuesCount = len(df.loc[df['message'].str.startswith("Error: Request failed with", na=False)])
            testDataIssuesCount = len(df.loc[df['message'].str.startswith("TEST_DATA_ERROR", na=False)])
            scriptingIssuesCount = (totalFailCount + blockedCount) - (
                orchestrationIssuesCount + labIssuesCount + appCrashIssuesCount + environmentIssuesCount + testDataIssuesCount
            )
            with open(failureListFileName, "a", encoding="utf-8") as myfile:
                myfile.write(str(error.strip())+'\n*******************************************\n')
    
    # Top 5 failure reasons
    topFailureDict = {}
    failureDict = Counter(cleanedFailureList)
    count_total = int(str(os.environ["recommendations"]))
    for commonError, commonErrorCount in failureDict.most_common(count_total):
        topFailureDict[commonError] = int(commonErrorCount)
    suggesstionsDict = {}
    # reach top errors and clean them
    i = 0
    for commonError, commonErrorCount in topFailureDict.items():
        if "ERROR: No device was found" in commonError:
            error = (
                "Raise a support case for the error: *|*"
                + commonError.strip() + "*|*"
            )
        elif "Cannot open device" in commonError:
            error = (
                "Reserve the device/ use perfecto lab auto selection feature to avoid the error:  *|*"
                + commonError.strip() + "*|*"
            )
        elif (
            '(UnknownError) Failed to execute command button-text click: Needle not found for expected value: "Allow" (java.lang.RuntimeException)'
            in commonError
        ):
            error = (
                "Allow text/popup was not displayed as expected. It could be an environment issue as the error: *|*"
                + commonError.strip() + "*|*"
            )
        else:
            error = (unicodedata.normalize("NFKD", str(commonError.strip())))
        report_link = df.loc[df['message'].str.startswith(str(commonError.strip()) , na=False), "reportURL"].iloc[0]
        suggesstionsDict[error] = [int(commonErrorCount), report_link]
    eDict = {}
    eDict = edict(
        {
            "status": [
                {
                    "#Total": "Count ->",
                    "#Executions": createHref(jobName, reportTag, '', totalTCCount),
                    "#Pass": createHref(jobName, reportTag, 'PASSED', totalPassCount),
                    "#Failed": createHref(jobName, reportTag, 'FAILED', totalFailCount),
                    "#Blocked": createHref(jobName, reportTag, 'BLOCKED', blockedCount),
                    "#Unknowns": createHref(jobName, reportTag, 'UNKNOWN', totalUnknownCount),
                    "Overall Pass %": str(
                        int(percentageCalculator(totalPassCount, totalTCCount))
                    )
                    + "%",
                },
            ],
            "issues": [
                {
                    "#Issues": "Count ->",
                    "#Scripting": scriptingIssuesCount,
                    "#App Crash": appCrashIssuesCount,
                    "#Environment Issues": environmentIssuesCount,
                    "#Test Data Issues": testDataIssuesCount,
                    "#Lab": labIssuesCount,
                    "#Orchestration": orchestrationIssuesCount,
                },
            ],
            "recommendation": [
            ],
        }
    )
    header = 'style="box-sizing: border-box; float: left; width: 100%; padding: 1px 0; text-align: center; cursor: pointer; font-size: 16px; color: black; background-color: darkkhaki; border: 3px solid antiquewhite;"'
    perJob = []
    tag = ''
    for i, job in enumerate(jobName.split(";")):
        if(job != ""):
            job_df = df.loc[(df["job/name"] == job)]
            job_total = job_df.shape[0]
            failed = job_df[(job_df["Test Status"] == "FAILED")].shape[0]
            passed = job_df[(job_df["Test Status"] == "PASSED")].shape[0]
            blocked = job_df[(job_df["Test Status"] == "BLOCKED")].shape[0]
            unknown = job_df[(job_df["Test Status"] == "UNKNOWN")].shape[0]
            perJob.append({ "Job": createHref(job, "", '', job), "#Executions": createHref(job, "", '', job_total), "#Pass": createHref(job, "", 'PASSED', passed) , "#Failed": createHref(job, "", 'FAILED', failed), "#Blocked": createHref(job, "", 'BLOCKED', blocked), "#Unknowns": createHref(job, "", 'UNKNOWN', unknown), "Overall Pass %": str(int(percentageCalculator(passed, job_total)))+ "%"})
   
    global per_job_status
    global per_job_status_email
    per_job_status = ""
    per_job_status_email = ""
    per_job = pandas.DataFrame.from_dict(perJob)

    if(per_job.shape[0] > 0):
        per_job =  per_job.to_html(
            classes="mystyle",
            table_id="report",
            index=False,
            render_links=True,
            escape=False,
        )
        per_job_status = """ <br><label for="tab2" style="background:  darkkhaki !important; color:rgb(15, 61, 16) !important;">Job wise Summary</label><br><br>""" + style_df_email(per_job)
        per_job_status_email = style_df_email(per_job)
        per_job_status_email = """ <br><div """ + header+ """><center><label for="tab2" style="background:  darkkhaki !important; color:black !important;font-weight:bold;">Job wise Summary</label></center></div><br><br>""" + per_job_status_email
       
    
    perTag = []
    tags_cols = [col for col in df.columns if 'tags' in col]
    exp_tags = ""
    if(reportTag == ''):
        exp_tags = os.environ["recommend_tag"]
    else:
        exp_tags = reportTag
    for i, tag in enumerate(exp_tags.split(";")):
        if(tag != ""):
            query = '' 
            for i, col in enumerate(tags_cols):
                query += '`' + col + "` == '" + tag + "'"
                if(i != (len(tags_cols) - 1)):
                    query += ' or '
            tag_df = df.query(query)
            job_total = tag_df.shape[0]
            failed = tag_df[(tag_df["Test Status"] == "FAILED")].shape[0]
            passed = tag_df[(tag_df["Test Status"] == "PASSED")].shape[0]
            blocked = tag_df[(tag_df["Test Status"] == "BLOCKED")].shape[0]
            unknown = tag_df[(tag_df["Test Status"] == "UNKNOWN")].shape[0]
            perTag.append({ "Tag": createHref(jobName, tag, '', tag), "#Executions": createHref(jobName, tag, '', job_total), "#Pass": createHref(jobName, tag, 'PASSED', passed) , "#Failed": createHref(jobName, tag, 'FAILED', failed), "#Blocked": createHref(jobName, tag, 'BLOCKED', blocked), "#Unknowns": createHref(jobName, tag, 'UNKNOWN', unknown), "Overall Pass %": str(int(percentageCalculator(passed, job_total)))+ "%"})
   
    global per_tag_status
    global per_tag_status_email
    per_tag_status = ""
    per_tag_status_email = ""
    per_tag = pandas.DataFrame.from_dict(perTag)

    if(per_tag.shape[0] > 0):
        per_tag = per_tag.to_html(
            classes="mystyle",
            table_id="report",
            index=False,
            render_links=True,
            escape=False,
        )
        per_tag_status =  """ <br><label for="tab2" style="background:  darkkhaki !important; color:rgb(15, 61, 16) !important;">Tag wise Summary</label><br><br>""" + style_df_email(per_tag)
        per_tag_status_email = style_df_email(per_tag)
        per_tag_status_email = """ <br><div """ + header+ """><center><label for="tab2" style="background:  darkkhaki !important; color:black !important;font-weight:bold;">Tag wise Summary</label></center></div><br><br>""" + per_tag_status_email
       
    recommendations_count = int(str(os.environ["recommendations"]))
    i = 0
    dynamic_rec = []
    while i < recommendations_count:
        dynamic_rec.append({ "Recommendations": "-", "Occurences":"-", "ReportURL":"-", "Rank": i+1, "impact": "0"})
        i += 1
    eDict['recommendation'] = dynamic_rec
    jsonObj = edict(eDict)
    
    if(bool):
        if float(percentageCalculator(totalUnknownCount, totalTCCount)) >= 30:
            suggesstionsDict[
                "# Fix the unknowns. The unknown script ratio is too high (%) : "
                + str(percentageCalculator(totalUnknownCount, totalTCCount))
                + "%"
            ] = [percentageCalculator(
                totalPassCount + totalUnknownCount, totalTCCount
            ) - percentageCalculator(
                totalPassCount, totalTCCount
            ), "-"]
        if len(suggesstionsDict) < count_total:
            if (topfailedTCNames.shape[0]) > 1:
                for tcName, status in topfailedTCNames.itertuples(index=False):
                    suggesstionsDict[
                        "# Fix the top failing tests listed under 'Top Failed Tests' "
                    ] = [1,"-"]
                    break

        if len(suggesstionsDict) < count_total:
            if int(percentageCalculator(totalFailCount, totalTCCount)) > 15:
                if totalTCCount > 0:
                    suggesstionsDict[
                        "# Fix the failures. The total failures % is too high (%) : "
                        + str(percentageCalculator(totalFailCount, totalTCCount))
                        + "%"
                    ] = [int(percentageCalculator(totalFailCount, totalTCCount)), "-"]
        if len(suggesstionsDict) < count_total:
            if float(percentageCalculator(totalPassCount, totalTCCount)) < 80 and (
                totalTCCount > 0
            ):
                suggesstionsDict[
                    "# Fix the failures. The total pass %  is too less (%) : "
                    + str(int(percentageCalculator(totalPassCount, totalTCCount)))
                    + "%"
                ] = [int(
                    (100
                    - (
                        percentageCalculator(
                            totalPassCount + totalUnknownCount, totalTCCount
                        )
                        - percentageCalculator(totalPassCount, totalTCCount)
                    )
                ) - int(percentageCalculator(totalPassCount, totalTCCount))),"-"]
    if len(suggesstionsDict) < count_total:
        if totalTCCount == 0:
            suggesstionsDict[
                "# There are no executions for today. Try Continuous Integration with any tools like Jenkins and schedule your jobs today. Please reach out to Professional Services team of Perfecto for any assistance :) !"
            ] = [100,"-"]
        elif int(percentageCalculator(totalPassCount, totalTCCount)) > 80:
            print(str(int(percentageCalculator(totalPassCount, totalTCCount))))
            suggesstionsDict["# Great automation progress. Keep it up!"] = [0, "-"]

        int(percentageCalculator(totalFailCount, totalTCCount)) > 15
    counter = 0
    totalImpact = 0
    for sugg, commonErrorCount in sorted(suggesstionsDict.items(), key = lambda x:(x[1],x[0]), reverse=True)[:count_total]:
        impact = 1
        if str(sugg) != "" and str(commonErrorCount[0]) !="":
            if sugg.startswith("# "):
                jsonObj.recommendation[counter].ReportURL =  '-'
                sugg = sugg.replace("# ", "")
                impact = str(int(float(str(commonErrorCount[0]))))
                jsonObj.recommendation[counter].Occurences = "-"
            else:
                jsonObj.recommendation[counter].Occurences =  int(float(str(commonErrorCount[0])))
                jsonObj.recommendation[counter].ReportURL =  '<a target="_blank" href="' + commonErrorCount[1] + '">link</a>'
                impact = str(percentageCalculator(
                    totalPassCount + int(float(str(commonErrorCount[0]))), totalTCCount
                ) - percentageCalculator(totalPassCount, totalTCCount))
            jsonObj.recommendation[counter].impact = (
                str(("%.2f" % round(int(float(str(impact))), 2))) + "%"
            )
            jsonObj.recommendation[counter].Recommendations = html.escape(
                sugg.replace("*|*", "'")
                .replace("{", "{{")
                .replace("}", "}}")
                .strip()
            )
            totalImpact += round(int(float(str(impact))), 2)
        counter += 1
    global execution_status
    execution_status = pandas.DataFrame.from_dict(jsonObj.status)
    execution_status = execution_status.to_html(
        classes="mystyle",
        table_id="report",
        index=False,
        render_links=True,
        escape=False,
    )
    execution_status = style_df_email(execution_status)
    global issues
    issues = ""
    global issues_email
    issues_email = ""
    if("true" in str(os.environ["showIssues"])):
        issues_df = pandas.DataFrame.from_dict(jsonObj.issues)
        issues_ori = issues_df.to_html(
            classes="mystyle",
            table_id="report",
            index=False,
            render_links=True,
            escape=False,
        )
        header = 'style="box-sizing: border-box; float: left; width: 100%; padding: 1px 0; text-align: center; cursor: pointer; font-size: 16px; color: black; background-color: darkkhaki; border: 3px solid antiquewhite;"'
        issues = """</div></div><input type="radio" id="tab3" name="tabs" checked=""/><label for="tab3">Issues</label><div class="tab-content1">
          <div class="reportDiv">""" + issues_ori
        issues = style_df_email(issues)
        tabcontent = 'style="box-sizing: border-box; padding: 10px; height: auto; -moz-transition: height 1s ease; -webkit-transition: height 1s ease; -o-transition: height 1s ease; transition: height 1s ease; overflow: auto; display: inline;justify-content: center;"'
        reportDiv = (
            'style="box-sizing: border-box; overflow-x: visible; text-align: -webkit-center;"'
        )
        issues_email = str("""<br></div><div """
        + header
        + """</div>"""
        + """<b><center>Issues</center></b></label></div><div class="tab-content1" """
        + tabcontent
        + """>
        <div class="reportDiv" """
        + reportDiv
        + """>""") + style_df_email(issues_ori)
    else:
        issues = """</div></div><input type="radio" id="tab3" name="tabs" checked=""/><label for="tab3">Groups</label><div class="tab-content1">
          <div class="reportDiv">""" + issues
    recommendations = pandas.DataFrame.from_dict(jsonObj.recommendation)
    # recommendations = recommendations.sort_values(['impact'], ascending=False)
    if totalImpact > 100:
        recommendations.columns = [
            "Issues",
            "Occurences",
            "ReportURL",
            "Rank",
            "Pass% Increase",
        ]
    else:
        recommendations.columns = [
            "Issues",
            "Occurences",
            "ReportURL",
            "Rank",
            "Pass% Increase - " + str(round(totalImpact, 2)) + "%",
        ]
    recommendations = recommendations[
        recommendations.Issues.astype(str) != "-"
    ]
    recommendations = recommendations.to_html(
        classes="mystyle",
        table_id="report",
        index=False,
        render_links=True,
        escape=False,
    )
    return style_df_email(recommendations)

        
def process_failures(df, failed_blocked, topfailedTCNames, graphs, interactive_graphs, failed, passed, blocked):
    
    global recommendations
    recommendations = get_recommendations(df, failed_blocked, topfailedTCNames, failed, passed, blocked, str(os.environ["cloudName"]), True)
    
    tags_cols = [col for col in df.columns if 'tags' in col]
    recommend_tag = os.environ["recommend_tag"]
    tagrec = ""
    tagLinks = ""
    if recommend_tag != "":
        for i, tag in enumerate(recommend_tag.split(";")):
            query = '' 
            for i, col in enumerate(tags_cols):
                query += '`' + col + "` == '" + tag + "'"
                if(i != (len(tags_cols) - 1)):
                    query += ' or '
            tag_df = df.query(query)
            tagrec += """ <br><div style="padding:1px 0; background-color:darkkhaki; color:black;text-align:center;font-family:Verdana,sans-serif;font-size:16px;width:100%;cursor:pointer;">
                        """+ tag + """ Recommendations
                </div>"""
            tagrec += get_recommendations(df, tag_df, topfailedTCNames, failed, passed, blocked, str(tag), False)
            tagLinks += """ &nbsp; <a href="./"""  + str(tag) + """_failures.html" style="white-space:nowrap;text-decoration:none;background-color: rgb(195, 132, 49);color: white;padding: 8px 16px;font-family: sans-serif;border-radius: 3px;">"""  + str(tag) + """ Failures</a>"""
            
    #prepares graphs & interactive graphs
    with open(email_report_filename, "a") as f:
        f.write(
            get_html_string_email(graphs, tagrec).format(
                table=df.to_html(
                    classes="mystyle",
                    table_id="report",
                    index=False,
                    render_links=True,
                    escape=False,
                )
            )
        )
    graphs.clear()
    with open(live_report_filename, "a") as f:
        f.write(
            get_html_string(graphs, tagrec, tagLinks).format(
                table=df.to_html(
                    classes="mystyle",
                    table_id="report",
                    index=False,
                    render_links=True,
                    escape=False,
                )
            )
        )
    with open(live_report_filename, "a") as f:
        f.write(''.join(interactive_graphs)
        )       

def main():
    """
    Runs the perfecto actions and reports
    """
    try:
        start_time = datetime.now().replace(microsecond=0)
        freeze_support()
        init()
        #     """fix Python SSL CERTIFICATE_VERIFY_FAILED"""
        if not os.environ.get("PYTHONHTTPSVERIFY", "") and getattr(
            ssl, "_create_unverified_context", None
        ):
            ssl._create_default_https_context = ssl._create_unverified_context
        parser = argparse.ArgumentParser(description="Perfecto Actions Reporter")
        parser.add_argument(
            "-c",
            "--cloud_name",
            metavar="cloud_name",
            help="Perfecto cloud name. (E.g. demo) or add it as a cloudName environment variable",
            nargs="?",
        )
        parser.add_argument(
            "-s",
            "--security_token",
            metavar="security_token",
            type=str,
            help="Perfecto Security Token/ Pass your Perfecto's username and password in user:password format  or add it as a securityToken environment variable",
            nargs="?",
        )
        parser.add_argument(
            "-d",
            "--device_list_parameters",
            metavar="device_list_parameters",
            type=str,
            help="Perfecto get device list API parameters to limit device list. Support all API capabilities which selects devices based on reg ex/strings,  Leave it empty to select all devices",
            nargs="?",
        )
        parser.add_argument(
            "-u",
            "--user_list_parameters",
            metavar="user_list_parameters",
            type=str,
            help="Perfecto get user list API parameters to limit user list. Support all API capabilities which selects users based on applicable parameters,  Leave it empty to select all users",
            nargs="?",
        )
        parser.add_argument(
            "-t",
            "--device_status",
            type=str,
            metavar="Different types of Device Connection status",
            help="Different types of Device Connection status. Values: all. This will showcase all the device status like Available, Disconnected, un-available & Busy. Note: Only Available devices will be shown by default",
            nargs="?",
        )
        parser.add_argument(
            "-a",
            "--actions",
            metavar="actions",
            type=str,
            help="Perfecto actions seperated by semi-colon. E.g. reboot:true;cleanup:true;get_network_settings:true",
            nargs="?",
        )
        parser.add_argument(
            "-r",
            "--refresh",
            type=str,
            metavar="refresh",
            help="Refreshes the page with latest device status as per provided interval in seconds",
            nargs="?",
        )
        parser.add_argument(
            "-o",
            "--output",
            type=str,
            metavar="output in html",
            help="output in html. Values: true/false. Default is true",
            nargs="?",
        )
        parser.add_argument(
            "-l",
            "--logo",
            type=str,
            metavar="shows customer logo",
            help="shows client logo if valid official client website url is specified in this sample format: www.perfecto.io",
            nargs="?",
        )
        parser.add_argument(
            "-e",
            "--email",
            type=str,
            metavar="prepares AI based emailable and live report along with statistics & recommendations",
            help="creates a downloadable csv/xlsx of reporting data along with AI emailable & live report with live charts, AI predictions and recommendations.",
            nargs="?",
        )
        parser.add_argument(
            "-b",
            "--bgcolor",
            type=str,
            metavar="sets the background color in report",
            help="overrides the background color in report based on provided hex color",
            nargs="?",
        )
        args = vars(parser.parse_args())
        try:
            if not args["cloud_name"]:
                print("Loading cloudName: " + os.environ["cloudName"] + " from environment variable.")
            else:
                os.environ["cloudName"] = args["cloud_name"]
        except Exception:
            if not args["cloud_name"]:
                parser.error(
                    "cloud_name parameter is empty. Either Pass the argument -c followed by cloud_name, eg. perfectoai -c demo or add it as a cloudName environment variable"
                )
                exit
            os.environ["cloudName"] = args["cloud_name"]
        try:
            if not args["security_token"]:
                print("Loading securityToken: " + os.environ["securityToken"] + " from environment variable.")
            else:
                os.environ["securityToken"] = args["security_token"]
        except Exception:
            if not args["security_token"]:
                parser.error(
                    "security_token parameter is empty. Pass the argument -c followed by cloud_name, eg. perfectoai -c demo -s <<TOKEN>> || perfectoai -c demo -s <<user>>:<<password>> or add it as a securityToken environment variable"
                )
                exit
            os.environ["securityToken"] = args["security_token"]
        if args["email"]:
            os.environ["bgcolor"] = "beige"
            if args["bgcolor"]:
                os.environ["bgcolor"] = args["bgcolor"]
            email_report = args["email"]

            try:
                global criteria
                global jobNumber
                global jobName
                global startDate
                global endDate
                global consolidate
                global trends
                global report
                global tags
                global reportTag
                global live_report_filename
                global recommend_tag
                global title_heading
                global showIssues
                consolidate = ""
                xlformat = "csv"
                port = ""
                orcaport = "8000"
                temp = ""
                regex = ""
                ci_name = ""
                ci_jenkins_url = ""
                ci_username = ""
                ci_token = ""
                recommendations_count = 5
                recommend_tag = ""
                report_array = email_report.split("|")
                title_heading = str(os.environ["cloudName"]).upper()  + " Report"
                os.environ["title"] = title_heading
                showIssues = "true"
                os.environ["showIssues"] = showIssues
                for item in report_array:
                    if "report" in item:
                        report, criteria = get_report_details(
                            item, temp, "report", criteria
                        )
                    if "jobName" in item:
                        jobName, criteria = get_report_details(
                            item, temp, "jobName", criteria
                        )
                    if "jobNumber" in item:
                        jobNumber, criteria = get_report_details(
                            item, temp, "jobNumber", criteria
                        )
                    if "startDate" in item:
                        startDate, criteria = get_report_details(
                            item, temp, "startDate", criteria
                        )
                    if "endDate" in item:
                        endDate, criteria = get_report_details(
                            item, temp, "endDate", criteria
                        )
                    if "consolidate" in item:
                        consolidate, criteria = get_report_details(
                            item, temp, "consolidate", criteria
                        )
                    if "xlformat" in item:
                        xlformat, criteria = get_report_details(
                            item, temp, "xlformat", criteria
                        )
                    if "port" in item:
                        port, criteria = get_report_details(
                            item, temp, "port", criteria
                        )
                    if "orcaport" in item:
                        orcaport, criteria = get_report_details(
                            item, temp, "orcaport", criteria
                        )
                    if "trends" in item:
                        trends, criteria = get_report_details(
                            item, temp, "trends", criteria
                        )
                    if "tags" in item:
                        tags, criteria = get_report_details(
                            item, temp, "tags", criteria
                        )
                    if "reportTag" in item:
                            reportTag, criteria = get_report_details(
                            item, temp, "reportTag", criteria
                        )
                    if "attachmentName" in item:
                        live_report_filename, criteria = get_report_details(
                            item, temp, "attachmentName", criteria
                        )
                    if "regex" in item:
                        regex, criteria = get_report_details(
                            item, temp, "regex", criteria
                        )
                    if "ci_name" in item:
                            ci_name, criteria = get_report_details(
                            item, temp, "ci_name", criteria
                        )
                    if "ci_jenkins_url" in item:
                            ci_jenkins_url, criteria = get_report_details(
                            item, temp, "ci_jenkins_url", criteria
                        )
                    if "ci_username" in item:
                            ci_username, criteria = get_report_details(
                            item, temp, "ci_username", criteria
                        )
                    if "ci_token" in item:
                            ci_token, criteria = get_report_details(
                            item, temp, "ci_token", criteria
                        )
                    if "recommendations" in item:
                            recommendations_count, criteria = get_report_details(
                            item, temp, "recommendations", criteria
                        )
                    if "recommend_tag" in item:
                                recommend_tag, criteria = get_report_details(
                            item, temp, "recommend_tag", criteria
                        )
                    if "title" in item:
                        title_heading, criteria = get_report_details(
                            item, temp, "title", criteria
                        )
                    if "showIssues" in item:
                        showIssues, criteria = get_report_details(
                        item, temp, "showIssues", criteria
                    )
                        
            except Exception as e:
                raise Exception(
                    "Verify parameters of report, split them by | seperator/ " + str(e)
                )
                sys.exit(-1)
            if "attachmentName=" in email_report:
                live_report_filename = live_report_filename + ".html"
            os.environ["title"] = title_heading
            os.environ["xlformat"] = xlformat
            os.environ["regex"] = ""
            os.environ["regex"] = regex
            os.environ["consolidate"] = ""
            os.environ["consolidate"] = consolidate
            os.environ["orcaport"] = orcaport
            os.environ["ci_name"] = ""
            os.environ["ci_name"] = ci_name
            os.environ["ci_jenkins_url"] = ""
            os.environ["ci_jenkins_url"] = ci_jenkins_url
            os.environ["ci_username"] = ""
            os.environ["ci_username"] = ci_username
            os.environ["ci_token"] = ""
            os.environ["ci_token"] = ci_token
            os.environ["recommendations"] = str(recommendations_count)
            os.environ["recommend_tag"] = ""
            os.environ["recommend_tag"] = str(recommend_tag)
            os.environ["showIssues"] = str(showIssues)
            
            filelist = glob.glob(os.path.join("*." + xlformat))
            for f in filelist:
                os.remove(f)
            filelist = glob.glob(os.path.join("*.txt"))
            for f in filelist:
                os.remove(f)
            filelist = glob.glob(os.path.join("*.html"))
            for f in filelist:
                os.remove(f)

            graphs, interactive_graphs, df = prepareReport(jobName, jobNumber, reportTag)
            if not jobName:
                if not reportTag:
                    criteria = "Start: " + startDate + ", End: " + endDate
            else:
                criteria += "Job: " + jobName.replace(";", "; ")
            if jobNumber != "":
                criteria += "<br> Build Number: " + jobNumber.replace(";", "; ")
            if os.environ["consolidate"] != "":
                criteria += (
                    ", start: "
                    + str(df["startTime"].iloc[-1]).split(" ", 1)[0]
                    + ", end: "
                    + str(df["startTime"].iloc[0]).split(" ", 1)[0]
                )
            elif startDate != "":
                criteria += " <br>Start: " + startDate + ", End: " + endDate
            global execution_summary
            global title
            title = ""
            if tags != "":
                title = report + criteria + "<br>Info: " + tags
            else:
                title = report + criteria
            if reportTag != "":
                title += "<br> tags: " + reportTag
            execution_summary = create_pie(df, "", "status", "device_summary",)
            failed = df[(df["status"] == "FAILED")]
            passed = df[(df["status"] == "PASSED")]
            blocked = df[(df["status"] == "BLOCKED")]
            failed_blocked = df[
                (df["status"] == "FAILED") | (df["status"] == "BLOCKED")
            ]
            
            # monthly stats
            df["platforms/0/deviceType"] = df["platforms/0/deviceType"].fillna("Others")
            df["platforms/0/os"] = df["platforms/0/os"].fillna("Others")
            df = df.rename(
                columns={
                    "platforms/0/deviceType": "Platform",
                    "platforms/0/os": "OS",
                    "status": "Test Status",
                    "failureReasonName": "Custom Failure Reason",
                    "platforms/0/osVersion": "Version"
                }
            )
            global monthlyStats
            monthlyStats = df.pivot_table(
                index=["month", "week", "Platform", "OS", "Version"],
                columns="Test Status",
                values="name",
                aggfunc="count",
                margins=True,
                fill_value=0,
            ).fillna("")
            for column in monthlyStats.columns:
                monthlyStats[column] = (
                    monthlyStats[column].astype(str).replace("\.0", "", regex=True)
                )
            monthlyStats = monthlyStats.to_html(
                classes="mystyle",
                table_id="report",
                index=True,
                render_links=True,
                escape=False,
            )
            monthlyStats = style_df_email(monthlyStats)
            global failurereasons
            if "Custom Failure Reason" not in df.columns:
                df["Custom Failure Reason"] = ""
            df.loc[(df["Custom Failure Reason"].astype(str) == "unclassified error") & (df["Test Status"].astype(str) == "PASSED"), "Custom Failure Reason"] = np.NaN
            failurereasons = pandas.crosstab(
                df["Custom Failure Reason"], df["Test Status"]
            )
            # print (failurereasons)
            failurereasons = failurereasons.to_html(
                classes="mystyle",
                table_id="report",
                index=True,
                render_links=True,
                escape=False,
            )
            failurereasons = style_df_email(failurereasons)
            # top failed TCs
            topfailedTCNames = (
                failed.groupby(["name"])
                .size()
                .reset_index(name="#Failed")
                .sort_values("#Failed", ascending=False)
                .head(5)
            )
            reportURLs = []
            for ind in topfailedTCNames.index:
                reportURLs.append(
                    failed.loc[
                        failed["name"] == topfailedTCNames["name"][ind], "reportURL"
                    ].iloc[0]
                )
            topfailedTCNames["Result"] = reportURLs
            topfailedTCNames["Result"] = topfailedTCNames["Result"].apply(
                lambda x: "{0}".format(x)
            )
            for ind in topfailedTCNames.index:
                topfailedTCNames.loc[topfailedTCNames["name"].index == ind, "name"] = (
                    '<a target="_blank" href="'
                    + topfailedTCNames["Result"][ind]
                    + '">'
                    + topfailedTCNames["name"][ind]
                    + "</a>"
                )
            topfailedTCNames = topfailedTCNames.drop("Result", 1)
            topfailedTCNames.columns = ["Top Failed Tests", "#Failed"]
            # print(str(topfailedTCNames))
            global topfailedtable
            topfailedtable = topfailedTCNames.to_html(
                classes="mystyle",
                table_id="report",
                index=False,
                render_links=True,
                escape=False,
            )
            topfailedtable = style_df_email(topfailedtable)
            process_failures(df, failed_blocked, topfailedTCNames, graphs, interactive_graphs, failed, passed, blocked)
            import http.server
            import socket
            import socketserver
            import webbrowser

            if port != "":
                PORT = int(port)
                Handler = http.server.SimpleHTTPRequestHandler
                url = (
                    "http://"
                    + socket.gethostbyname(socket.gethostname())
                    + ":"
                    + str(PORT)
                    + "/"
                    + live_report_filename
                )
                print("Live dashboard url: " + url)
                with socketserver.TCPServer(("", PORT), Handler) as httpd:
                    print("serving at port", PORT)
                    webbrowser.open(url, new=0)
                    httpd.serve_forever()
            else:
                webbrowser.open(
                    "file://" + os.path.join(os.getcwd(), live_report_filename), new=0
                )
                print(
                    "Interactive Report: file://"
                    + os.path.join(os.getcwd(), live_report_filename)
                )
                print(
                    "Emailable Report: file://"
                    + os.path.join(os.getcwd(), email_report_filename)
                )
                end = datetime.now().replace(microsecond=0)
                print("Total Time taken:" + str(end - start_time))
        else:
            os.environ["bgcolor"] = "black"
            if args["bgcolor"]:
                os.environ["bgcolor"] = args["bgcolor"]
            if args["device_list_parameters"]:
                device_list_parameters = args["device_list_parameters"]
            else:
                device_list_parameters = "All devices"
            os.environ["DEVICE_LIST_PARAMETERS"] = device_list_parameters
            os.environ["USER_LIST_PARAMETERS"] = "All users"
            if args["user_list_parameters"]:
                os.environ["USER_LIST_PARAMETERS"] = args["user_list_parameters"]
            os.environ[
                "perfecto_logo"
            ] = "https://logo.clearbit.com/www.perforce.com?size=120"
            if args["logo"]:
                if str("www.").lower() not in str(args["logo"]).lower():
                    raise Exception(
                        "Kindly provide valid client website url. Sample format: www.perfecto.io"
                    )
                else:
                    new_logo = "https://logo.clearbit.com/" + args["logo"] + "?size=120"
                    validate_logo(new_logo)
                    os.environ["company_logo"] = new_logo
            else:
                os.environ["company_logo"] = os.environ["perfecto_logo"]
            os.environ["GET_NETWORK_SETTINGS"] = "False"
            reboot = "False"
            cleanup = "False"
            start_execution = "False"
            clean_repo = "NA"
            if args["actions"]:
                if "get_network_settings:true" in args["actions"]:
                    os.environ["GET_NETWORK_SETTINGS"] = "True"
                if "reboot:true" in args["actions"]:
                    reboot = "True"
                if "cleanup:true" in args["actions"]:
                    cleanup = "True"
                if "clean_repo" in args["actions"]:
                    clean_repo = args["actions"]
                else:
                    os.environ["clean_repo"] = "NA"
            os.environ["clean_repo"] = clean_repo
            # manage repo:
            clean_repo = os.environ["clean_repo"]
            if "NA" != clean_repo:
                try:
                    clean_repo_var = clean_repo.split("|")
                    if ";" in str(clean_repo_var[4]):
                        day = str(clean_repo_var[4]).split(";")[0]
                    else:
                        day = str(clean_repo_var[4])
                    repo_html = deleteOlderFiles(
                        REPOSITORY_RESOURCE_TYPE,
                        clean_repo_var[1],
                        clean_repo_var[2],
                        clean_repo_var[3],
                        day,
                    )
                except Exception as e:
                    raise Exception(
                        "Exception: "
                        + str(e)
                    )
                    sys.exit(-1)
            os.environ["CLEANUP"] = cleanup
            os.environ["REBOOT"] = reboot
            if (
                "True" in os.environ["GET_NETWORK_SETTINGS"]
                or "True" in reboot
                or "True" in cleanup
            ):
                start_execution = "True"
            os.environ["START_EXECUTION"] = start_execution
            os.environ["PREPARE_ACTIONS_HTML"] = "true"
            if args["output"]:
                if "false" in str(args["output"]).lower():
                    os.environ["PREPARE_ACTIONS_HTML"] = "false"
            os.environ["perfecto_actions_refresh"] = "false"
            if args["refresh"]:
                if int(args["refresh"]) >= 0:
                    os.environ["perfecto_actions_refresh"] = args["refresh"]
            # create results path and files
            create_dir(os.path.join(TEMP_DIR, "results"), True)
            create_dir(os.path.join(TEMP_DIR, "repo_results"), True)
            create_dir(os.path.join(TEMP_DIR, "output"), True)
            # result = get_xml_to_xlsx(RESOURCE_TYPE, "list", 'get_devices_list.xlsx')
            # get device list to excel
            devlist = Pool(processes=1)
            try:
                result = devlist.apply_async(
                    get_xml_to_xlsx, [RESOURCE_TYPE, "list", "get_devices_list.xlsx"]
                )
            except Exception:
                devlist.close()
                print(traceback.format_exc())
                sys.exit(-1)
            # user_html = get_json_to_xlsx(RESOURCE_TYPE_USERS, "list", 'get_users_list.xlsx')
            userlist = Pool(processes=1)
            try:
                user_html = userlist.apply_async(
                    get_json_to_xlsx,
                    [RESOURCE_TYPE_USERS, "list", "get_users_list.xlsx"],
                ).get()
                userlist.close()
                userlist.terminate()
            except Exception:
                userlist.close()
                print(traceback.format_exc())
                sys.exit(-1)
            if args["device_status"]:
                #             may require for debug single threads
                #             get_list("list;connected;false;green;Available")
                #             get_list("list;connected;true;red;Busy")
                #             get_list("list;disconnected;;red;Disconnected")
                #             get_list("list;unavailable;;red;Un-available")
                get_dev_list = [
                    "list;connected;true;red;Busy",
                    "list;disconnected;;red;Disconnected",
                    "list;unavailable;;red;Un-available",
                    "list;connected;false;green;Available",
                ]
                try:
                    procs = []
                    for li in get_dev_list:
                        proc = Process(target=get_list, args=(str(li),))
                        procs.append(proc)
                        proc.start()
                    for proc in procs:
                        proc.join()
                    for proc in procs:
                        proc.terminate()
                except Exception:
                    proc.terminate()
                    print(traceback.format_exc())
                    sys.exit(-1)
            else:
                if not args["device_list_parameters"]:
                    os.environ["DEVICE_LIST_PARAMETERS"] = "Available Devices only"
                get_list("list;connected;false;green;Available")
            if "NA" != clean_repo:
                prepare_html(user_html, repo_html, day)
            else:
                prepare_html(user_html, "", "")
            end = datetime.now().replace(microsecond=0)
            print("Total Time taken:" + str(end - start_time))
            # Keeps refreshing page with expected arguments with a sleep of provided seconds
            while "false" not in os.environ["perfecto_actions_refresh"]:
                time.sleep(int(os.environ["perfecto_actions_refresh"]))
                main()
            devlist.close()
            devlist.terminate()

            try:
                if not platform.system() == "Darwin":
                    os.system("taskkill /f /im perfectoai.exe")
                if os.name == 'nt':
                    si = subprocess.STARTUPINFO()
                    si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                    subprocess.call('taskkill /F /IM Electron.exe', startupinfo=si)
            except:
                pass
    except Exception as e:
        raise Exception("Oops!", e)


if __name__ == "__main__":
    main()
    sys.exit()
