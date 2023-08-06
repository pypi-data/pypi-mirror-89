
# ---------------------------------------------------------------------------------------------------------------
# PyxlSQL project
# This program and library is licenced under the European Union Public Licence v1.2 (see LICENCE)
# developed by fabien.battini@gmail.com
# ---------------------------------------------------------------------------------------------------------------

import re
import os
import importlib
import inspect
import openpyxl
import openpyxl.styles
from PyxlSql.pyxlErrors import PyxlSqlColumnError, PyxlSqlSheetError, PyxlSqlError


# ---------------------------------------------------------------------------------------------------------------
# class NamedWS
# ---------------------------------------------------------------------------------------------------------------


class NamedWS:
    """
    Additional metadata for each named sheet in the NamedWB
    NB: Column names are *case sensitive* (NOT converted to lower case)
    """
    general_style = 'General'
    int_style = '#,##0'
    percent_style = '0.0%'
    euro_style = '_ * #,##0_) [$€-x-euro1]_ '
    euro2_style = '_ * #,##0.00_) [$€-x-euro1]_ '

    def __init__(self, workbook, sheet, column_name_row=1, font=None):
        sheet.auto_filter.ref = None    # to avoid issues when auto_filter is not correctly managed by OpenPyxl

        self.sheet = sheet
        self.father = workbook
        self.title = sheet.title
        self.full_name = self.father.filename + ':[' + self.title + ']'
        self.current_row = sheet.max_row + 1    # first free row
        self.column_name_row = column_name_row  # number of the row where the column name is stored
        self.columns = []                       # a list of columns
        self.column_names = {}                  # a dictionary: string --> Column number
        self.column_styles = {}                 # a dictionary: string --> style
        self.indexes = {}                       # a dictionary: index -->
        self.book_default_font = openpyxl.styles.Font(name='Century Gothic', size=11) if font is None else font

        max_col = sheet.max_column+1
        for i in range(1, max_col):
            col_name = sheet.cell(row=column_name_row, column=i).value
            if col_name is not None and col_name != "" and col_name != " ":
                col_name = str(col_name)  # str(), because could be an integer, typically a year .lower()
                self.column_names[col_name] = i
                self.columns.append(col_name)
                cell = sheet.cell(row=column_name_row + 1, column=i)
                if cell is not None:
                    self.column_styles[col_name] = cell.number_format

    def get_row_range(self):
        """returns the range of ACTIVE rows"""
        return range(self.column_name_row + 1, self.current_row)

    def get_start_of_range(self):
        return self.column_name_row + 1

    def find_column(self, column_name):
        """returns the column number from its name"""
        if column_name is None:
            raise PyxlSqlColumnError("find_column(None)", self.full_name)

        lc_column_name = str(column_name)  # .lower()
        if lc_column_name not in self.column_names:
            raise PyxlSqlColumnError(column_name, self.full_name)
        return self.column_names[lc_column_name]

    def get_max_column(self):
        # Returns the ID of the last column
        return len(self.column_names)+1

    def get_column_range(self):
        return range(1, self.get_max_column())

    @staticmethod
    def build_key(*args):
        """
        Creates the hash key for a list of descriptors,
        Typically (Column name, Row name)

        CAVEAT: all identifiers are NO MORE turned into lower case
        """
        key = ""
        for v in args:
            key += "" if v is None else (":" + str(v))  # str(v).lower())
        return key

    def get_cell(self, row_nb, column_name):
        col = self.find_column(column_name)
        if row_nb is None:
            raise PyxlSqlColumnError("Undefined",
                                     f"source:{str(inspect.currentframe().f_lineno)}" +
                                     f":{str(inspect.currentframe().f_code.co_filename)}")
        return self.sheet.cell(row=row_nb, column=col)

    def get_val(self, row_nb, column_name):
        cell = self.get_cell(row_nb, column_name)
        return None if cell.value is None or cell.value == "" else cell.value

    # UNUSED, for future PIVOT
    # def get_string(self, row_nb, column_name):
    #     cell = self.get_cell(row_nb, column_name)
    #     if cell is None or cell.value is None or cell.value == "None":
    #         return ""
    #     return str(cell.value)
    #
    # def get_float(self, row_nb, column_name):
    #     cell = self.get_cell(row_nb, column_name)
    #     if cell is None:
    #         return 0
    #     val = cell.value
    #     if val is None or val == "":
    #         return 0
    #
    #     if isinstance(val, str) and val.startswith("="):
    #         val = val[1:]  # remove leading =
    #         int_val = eval(val)  # a tentative to manage simple cells, such as '=12.56-24.9'
    #         # print("EVAL", val, intVal)
    #         # TODO: Track Errors
    #         return int_val
    #     return float(val)  # ********** CAVEAT! must be Float, because can be less than 1
    #
    # def get_style(self, column_name):
    #     lc_column_name = str(column_name)  # .lower()
    #     if lc_column_name not in self.column_names:
    #         raise PyxlSqlColumnError(column_name, self.full_name)
    #     return self.column_styles[lc_column_name]
    #
    # def get_index(self, *column_names):
    #     """
    #     :param self: a sheet
    #     :param column_names: the list of columns to be indexed
    #     :return: the index, i.e. the dictionary (value of the indexed columns) --> row number
    #
    #     if the index was not yet created, creates it
    #     """
    #     key = NamedWS.build_key(*column_names)  # Manages SEVERAL columns
    #     if key not in self.indexes:
    #         index_hash = {}
    #         for row in self.get_row_range():
    #             values = []
    #             for col in column_names:
    #                 values.append(self.get_val(row, col))
    #             val_key = NamedWS.build_key(*values)
    #             if val_key not in index_hash:
    #                 index_hash[val_key] = []
    #             index_hash[val_key].append(row)
    #         self.indexes[key] = index_hash
    #     return self.indexes[key]

    # ------------ set methods

    def set_value(self, row_nb: int, column_name: str, value, number_format=None) -> None:
        cell = self.get_cell(row_nb, column_name)
        assert cell is not None

        if isinstance(value, str):
            value = re.sub("^ *", "", value)
            value = re.sub(" *$", "", value)

        cell.value = value
        # cell.font = self.book_default_font
        # number_format = number_format if number_format else self.get_style(column_name)

        if number_format:
            cell.number_format = number_format

        self.current_row = max(self.current_row, row_nb+1)  # current_row is the first FREE cell

    # UNUSED, for future PIVOT
    # def create_row(self, init):
    #     for column, value in init.items():
    #         self.set_value(self.current_row, column, value)
    #
    #     self.current_row += 1
    #     return self.current_row - 1
# ---------------------------------------------------------------------------------------------------
# class NamedWB
# ---------------------------------------------------------------------------------------------------


class NamedWB:
    """
    class to manage excel workbooks structured as a database
    """
    def __init__(self, file_name, create=False, column_name_row=1, font=None, file_path=None):
        self.file_path = file_path or []
        self.filename = self.find_file(file_name)
        local_path = os.path.dirname(self.filename)
        if local_path not in self.file_path:
            self.file_path.append(local_path)

        if self.filename is None:
            print(f"FATAL ERROR: Cannot find '{file_name}' in path '{self.file_path}', aborting")
            exit(-1)

        try:
            self.wb: openpyxl.workbook = openpyxl.load_workbook(filename=self.filename)
        except OSError as error:
            print(f"FATAL ERROR: Cannot open '{file_name}' : {str(error)}, aborting")
            exit(-1)

        self.sheets = {}  # a dictionary  string --> NamedWS
        self.book_column_name_row = column_name_row  # BookColumnNameRow  column_name_row
        self.book_default_font = openpyxl.styles.Font(name='Century Gothic', size=11) if font is None else font
        self.wbs = {}       # all workbooks referenced by this one
        self.imported = {}  # dictionary of imported symbols, will be used when eval is called
        self.sheets_to_delete = []

        if create:
            self.wb = openpyxl.Workbook()
            return

        for sheet in self.wb:
            self.sheets[sheet.title] = NamedWS(self, sheet, column_name_row=column_name_row)

    def find_file(self, filename):
        if filename is None:
            return None
        if os.path.exists(filename):
            return filename
        for d in self.file_path:
            if d[-1] != "/":
                d += "/"
            f = os.path.realpath(d + filename)
            if os.path.exists(f):
                return f
        return None

    def get_sheet(self, sheet_name, raise_except=True):
        if sheet_name is None:
            return None
        if sheet_name in self.sheets:
            return self.sheets[sheet_name]

        m = re.search("\"?\'?([^\"\'[]+)\"?\'?(\\[(.*)])?", sheet_name)
        if m is not None:
            first_name: str = m.group(1)
            second_name: str = m.group(3)
            if first_name is not None and second_name is not None:
                workbook = self.get_workbook(first_name)
                if workbook is not None and second_name in workbook.sheets:
                    return workbook.sheets[second_name]
            # here, we have not found the sheet

        if raise_except:
            raise PyxlSqlSheetError(sheet_name, "workbook")
        return None

    def delete_sheet(self, sheet_name):
        """Deletes the sheet from the workbook"""
        sheet = self.get_sheet(sheet_name)
        if sheet is not None:
            del self.wb[sheet_name]

    def save(self, file_name):
        current_dir = os.path.dirname(os.path.realpath(self.filename))
        self.wb.save(current_dir + "\\" + file_name)

    def get_workbook(self, name):
        if name not in self.wbs:
            self.wbs[name] = NamedWB(name, file_path=self.file_path)
        return self.wbs[name]

    def import_module(self, module, sub_modules=None):
        mod = importlib.import_module(module)
        if sub_modules is None:
            self.imported[module] = mod
            return

        if sub_modules == '*':
            if hasattr(mod, '__all__'):
                item_list = mod.__all__
            else:
                raise PyxlSqlError(f"ERROR: IMPORT {module} SUBS {sub_modules} : ABORTING the import",
                                   "        The module does not contain a __all__ symbol that allows importing *")
        else:
            item_list = sub_modules.split()

        for item in item_list:
            self.imported[item] = getattr(mod, item)
