from os.path import isfile
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.cell import Cell
from . import log, version
from .utils import is_vertical_range
from .parser import Parser

import re
import yaml
import dictdiffer

ANON_PRFX = 'anon_'

def useless_range(rng):
    """
    Check if rng is valid and useful range.
    Empty range is not significative.

    :param rng:
    :return: True | False
    """
    if not rng:
        return True


class Collector:
    """
    Main Collector class
    """
    label_names = [re.compile('^varname', re.IGNORECASE),
                   re.compile('^label', re.IGNORECASE)]
    model_names = [re.compile('^model', re.IGNORECASE)]
    data_names = [re.compile('^data', re.IGNORECASE)]

    def __init__(self, url:str, only_data:bool=False, relative:bool=False, add_fingerprint:bool=False,
                 parsed:bool=False):
        """
        Data injestion, we need 2 instances of the sheet:
          - wb_data: with static data
          - wb: with the formulas

        :param url: url to an input excel file
        :param only_data: read only static values
        :param parsed: use parsed formaulas instead of excel version
        :param add_fingerprint: add some punctual information in the internal representation
        :param relative: all area are treated as if they starts from Row1 Col1
        """
        if not isfile(url):
            raise FileNotFoundError("File {} does not exists".format(url))

        self.sheets = {}
        self.pseudo = {}
        self.url = url
        self.anon_models = {}
        self.parsed = parsed

        self.labels_as_data = True
        self.models_as_data = only_data
        log.debug("Labels read as {}".format( 'data' if self.labels_as_data else 'formula'))
        log.debug("Models read as {}".format('data' if self.models_as_data else 'formula'))
        if self.parsed:
            log.info("Parser enabled")

        # workbook injestion, we need 2 instances: data and formulas
        # if only_data is true each point to the same instance.

        self.wb_data = load_workbook(filename=self.url, data_only=True)
        if only_data:
            self.wb = self.wb_data
        else:
            self.wb = load_workbook(filename=self.url)
        self.relative = relative
        self.add_fingerprint = add_fingerprint

        self.sheetnames = self.wb_data.sheetnames
        self.named_ranges = self.wb_data.defined_names.definedName
        self.set_ranges()

        self.labels = self.handle_range(
            self.label_names,
            self.labels_as_data)

        # Here we have collected only labels, so if parsed is True
        # we need a data structure for bind position of formula to his label
        self.sheet_is_vertical = {}
        if self.parsed:
            # models pre-scan, this solves the anonymous models problems.
            # we don't know if they are vertical or horizontal
            self.models = self.handle_range(
                self.model_names,
                True)

            self.find_anonymous_models()

            self.pos_to_label = defaultdict(lambda: {})
            for sheet,rng in self.labels.items():
                if sheet not in self.sheet_is_vertical:
                    is_vert = is_vertical_range(rng.keys())
                    self.sheet_is_vertical[sheet] = is_vert
                else:
                    is_vert = self.sheet_is_vertical[sheet]
                for coord, label in rng.items():
                    row, col = coordinate_to_tuple(coord)
                    self.pos_to_label[sheet][col if is_vert else row] = label

            self.parser = Parser(collector=self)

            self.models = self.handle_range(
                self.model_names,
                self.models_as_data)
        else:
            self.models = self.handle_range(
                            self.model_names,
                            self.models_as_data)

            self.find_anonymous_models()

        self.data = self.handle_range(
                      self.data_names,
                      use_data=True)

        self.set_pseudo_excel()



    def find_anonymous_models(self):
        """
        Models can be defined without labels, so in this cases we must do
        more to:
          - Assign default labels
          - Understand if model is vertical or horizontal

        :return:
        """
        # Check for anonymous models
        for k,v in self.models.items():
            if k not in self.labels:
                self.sheet_is_vertical[k] = is_vertical_range(v)
                anon_labels = {f'{x}': f'{ANON_PRFX}{n+1}' for n, x in enumerate(v)}
                self.anon_models[k] = '{} anon labels assigned'.format(len(anon_labels))
                self.labels[k] = anon_labels

        for k,v in self.anon_models.items():
            log.warning("Found anonymous model {} : {}".format(k,v))


    def set_ranges(self):
        """
        Iter over named ranges and collect ALL ranges into:
            - self.ranges
            - self.params
            - self.text_ranges

        here we dont know if ranges are used, but we cannot distinguish cell_alias, and other stuff.

        :return:
            None
        """
        self.ranges = {}
        self.text_ranges = {}
        self.params = {}

        for x in self.named_ranges:
            for sheet, rng in x.destinations:
                if x.type == 'TEXT':
                    self.text_ranges[x.name] = x.value
                elif x.type == 'NUMBER':
                    self.params[x.name] = x.value
                else:
                    if useless_range(rng):
                        log.debug("{} range [{}]in sheet {} discarded ".format(x.name,rng,sheet))
                        continue
                    try:
                        cells = self.wb_data[sheet][rng]
                    except KeyError as err:
                        log.error(err)
                        continue

                    if isinstance(cells,Cell):
                        # Single named cell
                        self.ranges[x.name] = (cells,)
                    else:
                        # Range named
                        self.ranges[x.name] = sum(cells, ())

        log.debug("{} named ranges collected".format(len(self.ranges)))
        log.debug("{} parameters collected".format(len(self.params)))
        log.debug("{} text range collected".format(len(self.text_ranges)))

    def to_relative(self, rng_name:str, cell:Cell) -> str:
        """
        given a range name and a cell inside it, it return che cell position
        as R<row>C<col> as relative to the cell 0,0

        :param rng_name: collection of cells
        :param cell: target cell
        :return: relative position of the cell.
        """
        ref_cell = self.ranges[rng_name][0]
        min_row, min_col = ref_cell.row, ref_cell.column
        return "R{0}C{0}".format(cell.row - min_row + 1, cell.column - min_col + 1)


    def handle_range(self, labels: list, use_data:bool):
        """
        iter over collected ranges and restructure all data in a confortable nested dict.
        coll[<sheet_name>] = {..}

        :param labels: list of RE to use to undestand type of range
        :param use_data: Use data or formulas.
        :return: collection of all collected cells
        """
        coll = {}
        for lbl in self.ranges:
            for check_valid in labels:
                if check_valid.match(lbl):
                    cells = self.ranges[lbl]
                    sheet_names = set([x.parent.title for x in cells])
                    if len(sheet_names) > 1:
                        raise NotImplementedError("Range defined on multiple sheets is not handled")
                    sheet_name = sheet_names.pop()
                    if use_data:
                        coll[sheet_name] = {self.to_relative(lbl,x)
                                            if self.relative
                                            else x.coordinate:x.value
                                            for x in cells}
                    else:
                        coll[sheet_name] = {self.to_relative(lbl,x)
                                            if self.relative
                                            else x.coordinate:self.load_formula(sheet_name,  x.coordinate,  self.wb[sheet_name][x.coordinate].value)
                                            for x in cells}
                    break
        return coll


    def load_formula(self, sheet:str, position:str, s:str) -> str:
        """
        In the case we need to handle a formula we need an optional parameter
        to activate the parser. *self.parsed*
        If parsed is True so we do a further pass for rewrite in pythonic way.

        :param sheet:  sheet containing cell
        :param coordinate: cell coordinate
        :param s: input formula a simple string
        :return: s | transliterated version
        """
        if (not self.parsed) or (not isinstance(s,str)):
            return s

        log.debug('Parsing {} {} {}'.format(sheet, position, s))
        # In case of anonymous sheet we must force
        self.parser.set_current(sheet, coordinate_to_tuple(position))
        ret = self.parser.transform(s)
        log.debug('Parsed {}'.format(ret))
        return ret



    def set_pseudo_excel(self):
        if not self.pseudo:
            # Remember dict are ordered!
            if self.add_fingerprint:
                self.pseudo['xltoy.version'] = version
                self.pseudo['xltoy.filename'] = self.url
                self.pseudo['xltoy.datetime'] = datetime.now().isoformat()

            for sheet, labels in self.labels.items():
                self.pseudo[sheet] = dict(zip(labels.values(), self.models[sheet].values()))
                if sheet in self.data:
                    self.pseudo[sheet]['data'] = self.data[sheet]

            # We must handle data ranges not touched before, here
            # we have only data range not in sheet with labels
            for sheet in set(self.data) - set(self.labels):
                self.pseudo[sheet] = self.data[sheet]

    def to_yaml(self):
        self.set_pseudo_excel()
        return yaml.dump(self.pseudo)


class YamlCollector(Collector):
    """
    Collector for yaml format, it is used only to set
    self.pseudo attribute.
    """
    def __init__(self, url):
        if not isfile(url):
            raise FileNotFoundError("File {} does not exists".format(url))

        with open(url) as fin:
            self.pseudo = yaml.load(fin, Loader=yaml.FullLoader)



class DiffCollector:
    def __init__(self, url1, url2, only_data:bool=False, relative:bool=False, add_fingerprint:bool=False,
                 parsed:bool=False):
        """
        workbook differ, given 2 files (excel, yaml or json) it can do intelligent comparison

        :param url1:
        :param url2:
        :param only_data: ignore formulas and compare only values
        :param relative: all area are treated as if they starts from Row1 Col1
        :param parsed: use parsed formaulas instead of excel version
        :param add_fingerprint:
        """
        c1,c2 = [YamlCollector(url)
                 if url.lower().endswith('yaml')
                 else Collector(url,only_data=only_data, relative=relative,
                                    add_fingerprint=add_fingerprint, parsed=parsed)
                 for url in [url1,url2]]

        self.iter_differs = dictdiffer.diff(c1.pseudo,c2.pseudo)
        self.do_diff()

    def do_diff(self):
        self.diff = {}
        for kind, mid, sh_cells in self.iter_differs:
            if kind not in self.diff:
                self.diff[kind] = {}
            if isinstance(mid, list):
                if len(mid) == 1:
                    mid = mid[0]
                else:
                    raise RuntimeError("{} not understood".format(mid))
            if kind == 'change':
                sheet, label = mid.split('.')
                if sheet not in self.diff[kind]:
                    self.diff[kind][sheet] = {}
                self.diff[kind][sheet][label] = ' -> '.join(['{}'.format(x) for x in sh_cells])
            else:
                for sheet, cells in sh_cells:
                    if sheet not in self.diff[kind]:
                        self.diff[kind][sheet] = {}
                    self.diff[kind][sheet] = cells

    def to_yaml(self):
        if self.diff:
            print(yaml.dump(self.diff))

