from __future__ import annotations
from bs4 import Tag
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Union
from . import (
    BaseSimdakPaud,
    JENIS_KOMPONEN,
    JENIS_PENGGUNAAN,
    PENGGUNAAN,
    get_key,
    get_fuzz,
)

INDEX = "A"
MAPPING = {
    "jenis_komponen_id": "B",
    "jenis_penggunaan_id": "C",
    "jenisbelanja": "D",
    "qty": "E",
    "satuan": "F",
    "hargasatuan": "G",
    "data_id": "H",
}


@dataclass
class Rab(BaseSimdakPaud):
    jenis_komponen_id: int
    jenis_penggunaan_id: int
    jenisbelanja: str
    qty: int
    satuan: str
    hargasatuan: int
    data_id: str = ""

    def __post_init__(self) -> None:
        self.jenis_komponen_id = int(self.jenis_komponen_id)
        self.jenis_penggunaan_id = int(self.jenis_penggunaan_id)
        self.qty = int(self.qty)
        self.hargasatuan = int(self.hargasatuan)
        data_id = self.data_id
        if data_id:
            if "=" in data_id:
                data_id = data_id.split("=")[-1]
            elif "/" in data_id:
                data_id = data_id.split("/")[-1]
        self.data_id = data_id
        self._logger.debug(f"RPD [{self}]")

    def update(self, **kwargs) -> Rab:
        for k, v in kwargs.items():
            if hasattr(self, k):
                setattr(self, k, v)
        data = self.as_data(yt0="Simpan")
        url = self._base_url + f"boppaudrkas/update/id/{self.data_id}"
        res = self._session.post(url, data)
        if not res.ok:
            self._logger.warning(f"Gagal mengupdate data [{self.data_id}]")
        return self

    def delete(self) -> bool:
        url = self._base_url + f"boppaudrkas/delete/id/{self.data_id}"
        res = self._session.post(url, data=None)
        return res.ok

    def as_data(self, **kwargs) -> dict:
        if self.jenis_komponen_id not in JENIS_PENGGUNAAN:
            raise ValueError(f"jenis komponen {self.jenis_komponen_id} tidak valid!")
        if self.jenis_penggunaan_id not in JENIS_PENGGUNAAN[self.jenis_komponen_id]:
            raise ValueError(
                f"jenis penggunaan {self.jenis_penggunaan_id} tidak valid!"
            )
        data = {
            "Boppaudrkas[jenis_komponen_id]": self.jenis_komponen_id,
            "Boppaudrkas[jenis_penggunaan_id]": self.jenis_penggunaan_id,
            "Boppaudrkas[jenisbelanja]": self.jenisbelanja,
            "Boppaudrkas[qty]": self.qty,
            "Boppaudrkas[satuan]": self.satuan,
            "Boppaudrkas[hargasatuan]": self.hargasatuan,
        }
        data.update(kwargs)
        return data

    def as_dict(self, **kwargs) -> dict:
        return {
            "jenis_komponen_id": self.jenis_komponen_id,
            "jenis_penggunaan_id": self.jenis_penggunaan_id,
            "jenisbelanja": self.jenisbelanja,
            "qty": self.qty,
            "satuan": self.satuan,
            "hargasatuan": self.hargasatuan,
            "data_id": self.data_id,
        }

    def to_row(self, ws: Union[Worksheet, Workbook], row: int):
        data = self.as_dict()
        try:
            for k, v in data.items():
                col = MAPPING[k]
                ws[f"{col}{row}"] = v
            # TODO : Buat ini dynamic
            ws[f"B{row}"] = JENIS_KOMPONEN.get(self.jenis_komponen_id)
            ws[f"C{row}"] = PENGGUNAAN.get(self.jenis_penggunaan_id)
        except Exception as e:
            self._logger.exception(
                f"Gagal memasukkan data ke baris {row-1}, karena {e}"
            )

    def __eq__(self, other: object) -> bool:
        if not isinstance(other, Rab):
            return super().__eq__(other)
        return (
            self.jenis_komponen_id == other.jenis_komponen_id
            and self.jenis_penggunaan_id == other.jenis_penggunaan_id
            and self.jenisbelanja == other.jenisbelanja
            and self.qty == other.qty
            and self.satuan == other.satuan
            and self.hargasatuan == other.hargasatuan
        )

    def __str__(self) -> str:
        s = [
            f"JK: {JENIS_KOMPONEN[self.jenis_komponen_id]}",
            f"JP: {PENGGUNAAN[self.jenis_penggunaan_id]}",
            f"JB: {self.jenisbelanja}",
            f"Jumlah: {self.qty} {self.satuan}",
            f"Harga: {self.hargasatuan}",
        ]
        return ", ".join(s)

    @classmethod
    def from_tr(cls, tr: Tag) -> Rab:
        tds = tr.findAll("td")
        if not tds:
            raise ValueError("Data tidak ditemukan")
        jenis_komponen_id = get_key(tds[2].get_text(), JENIS_KOMPONEN)
        if not jenis_komponen_id:
            raise ValueError(f"Jenis Komponen {jenis_komponen_id} tidak valid")
        jenis_penggunaan_id = get_key(
            tds[3].get_text(), JENIS_PENGGUNAAN[jenis_komponen_id]
        )
        if not jenis_penggunaan_id:
            raise ValueError(f"Jenis Penggunaan {jenis_penggunaan_id} tidak valid")
        return cls(
            jenis_komponen_id=jenis_komponen_id,
            jenis_penggunaan_id=jenis_penggunaan_id,
            jenisbelanja=tds[4].get_text(),
            qty=int(tds[5].get_text()),
            satuan=tds[6].get_text(),
            hargasatuan=int(tds[7].get_text()),
            data_id=tds[10].find("a")["href"] if len(tds) == 11 else "",
        )

    @classmethod
    def from_row(cls, ws: Union[Worksheet, Workbook], row: int) -> Rab:
        data = {}
        for k, v in MAPPING.items():
            data[k] = ws[f"{v}{row}"].value
        col = MAPPING["jenis_komponen_id"]
        val = ws[f"{col}{row}"].value
        data["jenis_komponen_id"] = get_fuzz(val, JENIS_KOMPONEN, 1)
        col = MAPPING["jenis_penggunaan_id"]
        val = ws[f"{col}{row}"].value
        data["jenis_penggunaan_id"] = get_fuzz(val, PENGGUNAAN, 21)
        return cls(**data)

    @classmethod
    def from_table(cls, table: Tag) -> List[Rab]:
        results: List[Rab] = []
        for tr in table.findAll("tr"):
            try:
                result = cls.from_tr(tr)
                results.append(result)
            except ValueError as e:
                cls._logger.exception(e)
                continue
        cls._logger.info(f"Berhasil mendapatkan {len(results)} rpd")
        return results
